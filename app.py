import streamlit as st
import pandas as pd
import networkx as nx
import random
import io
import json
import re
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Camp Group Creator", layout="wide")
st.sidebar.title("🏕️ Camp Group Creator")

# =========================================================================================
# ========================= GLOBAL DATA PROCESSING (SHARED BY ALL PAGES) ==================
# =========================================================================================
st.sidebar.header("📂 1. Core Data Upload")
responses_file = st.sidebar.file_uploader("Preference Survey (CSV)", type=["csv"])
students_file  = st.sidebar.file_uploader("Student List (CSV)",      type=["csv"])

# ── Tool selection comes FIRST so Y8 processing can be gated by page ─────────
st.sidebar.markdown("---")
st.sidebar.header("🎯 2. Select Tool")
page = st.sidebar.radio("Select tool", ["🏕️ Y8 Group Creator", "🏔️ Y9 Journey Groups", "📋 Final Roster & Leader Builder", "🔍 Y9 Free Text Analyser"], label_visibility="collapsed")

# ── Y8 global data processing (only when a Y8 tool is active) ────────────────
df_merged_full = None
df_stud_pub = None
mtb_ability_col = None
all_students_list = []

if responses_file and students_file and page in ["🏕️ Y8 Group Creator", "📋 Final Roster & Leader Builder"]:
    responses_file.seek(0)
    students_file.seek(0)
    df_resp = pd.read_csv(responses_file)
    df_stud = pd.read_csv(students_file)

    # Clean data
    df_resp['Email address'] = df_resp['Email address'].astype(str).str.strip().str.lower()
    df_stud['Email'] = df_stud['Email'].astype(str).str.strip().str.lower()
    df_stud['Official Name'] = df_stud['Preferred name'].astype(str).str.strip() + " " + df_stud['Surname'].astype(str).str.strip()

    if 'Gender' in df_stud.columns: df_stud['Gender'] = df_stud['Gender'].astype(str).str.strip().str.lower()
    else: df_stud['Gender'] = 'o' 

    # Explicitly map the "Code" column to "Student ID"
    if 'Code' in df_stud.columns:
        df_stud['Student ID'] = df_stud['Code'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    else:
        id_col = next((c for c in df_stud.columns if 'id' in str(c).lower() and 'email' not in str(c).lower()), None)
        if id_col: df_stud['Student ID'] = df_stud[id_col].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        else: df_stud['Student ID'] = "N/A"

    df_stud_pub = df_stud.copy()

    df_merged_full = pd.merge(df_stud[['Email', 'Official Name', 'Rollgroup', 'Gender', 'Student ID']], 
                         df_resp, left_on='Email', right_on='Email address', how='left')

    df_merged_full['Responded'] = df_merged_full['Email address'].notna()

    valid_classes = df_resp['Which Connections class are you in? '].dropna().unique()
    rg_to_class = {f"8{str(c).strip()[0].upper()}": str(c).strip() for c in valid_classes}
    df_merged_full['Which Connections class are you in? '] = df_merged_full['Which Connections class are you in? '].fillna(
        df_merged_full['Rollgroup'].astype(str).str.strip().str.upper().map(rg_to_class)
    )
    df_merged_full = df_merged_full[df_merged_full['Which Connections class are you in? '].isin(valid_classes)]

    mtb_ability_col = next((c for c in df_merged_full.columns if 'comfortable are you riding a bike' in str(c).lower()), None)
    all_students_list = sorted(df_merged_full['Official Name'].dropna().unique().tolist())


# =========================================================================================
# ========================= PAGE 1: Y8 GROUP CREATOR ======================================
# =========================================================================================
if page == "🏕️ Y8 Group Creator":
    st.title("🏕️ Year 8 Camp Group Creator")

    # Default AI scoring weights
    default_weights = {
        "penalty_group_size_diff_multiplier": 10, "penalty_group_over_14": 500000, "penalty_forced_location": 1000000,
        "penalty_must_go_with": 1000000, "penalty_sole_gender": 200000, "penalty_zero_friends": 100000,
        "penalty_minority_gender_no_friends": 100000, "penalty_separation": 100000, "penalty_veto_activity": 50000,
        "penalty_minority_gender_with_friends": 1000, "reward_friend_1": 120, "reward_friend_2": 100, 
        "reward_activity_score_multiplier": 100
    }

    # Initialize session states
    if 'separation_pairs' not in st.session_state: st.session_state.separation_pairs = []
    if 'must_go_pairs' not in st.session_state: st.session_state.must_go_pairs = []
    if 'forced_locations' not in st.session_state: st.session_state.forced_locations = {}
    if 'not_attending' not in st.session_state: st.session_state.not_attending = []
    if 'generation_seed' not in st.session_state: st.session_state.generation_seed = 0
    if 'weights' not in st.session_state: st.session_state.weights = default_weights.copy()
    if 'optimized_groups' not in st.session_state: st.session_state.optimized_groups = {}
    if 'class_opt_states' not in st.session_state: st.session_state.class_opt_states = {}

    if df_merged_full is not None:
        
        # --- LOAD SETTINGS FEATURE ---
        st.sidebar.markdown("---")
        st.sidebar.header("📂 Load Rules & Weights")
        
        settings_file = st.sidebar.file_uploader("Load Saved JSON", type=["json"])
        
        if settings_file is not None:
            if st.session_state.get('last_loaded_file') != settings_file.file_id:
                try:
                    settings = json.load(settings_file)
                    st.session_state.separation_pairs = [tuple(x) for x in settings.get('separation_pairs', [])]
                    st.session_state.must_go_pairs = [tuple(x) for x in settings.get('must_go_pairs', [])]
                    st.session_state.forced_locations = settings.get('forced_locations', {})
                    
                    _all_stud_set = set(all_students_list)
                    _lower_stud_map = {n.lower(): n for n in all_students_list}
                    loaded_na = []
                    for _nm in settings.get('not_attending', []):
                        _nm_s = str(_nm).strip()
                        if _nm_s in _all_stud_set:
                            loaded_na.append(_nm_s)
                        elif _nm_s.lower() in _lower_stud_map:
                            loaded_na.append(_lower_stud_map[_nm_s.lower()])
                    st.session_state.not_attending = loaded_na
                    
                    if 'weights' in settings:
                        st.session_state.weights.update(settings['weights'])
                        if 'reward_friend_1' not in st.session_state.weights: st.session_state.weights['reward_friend_1'] = 120
                        if 'reward_friend_2' not in st.session_state.weights: st.session_state.weights['reward_friend_2'] = 100
                    
                    # Write directly into the widget key — identical to manual selection
                    st.session_state['y8_na_ms'] = loaded_na
                    st.session_state.not_attending = loaded_na
                    st.session_state.last_loaded_file = settings_file.file_id
                    st.sidebar.success("Settings loaded successfully!")
                    st.rerun() 
                except Exception as e:
                    st.sidebar.error("Error loading the rules file.")

        st.sidebar.markdown("---")

        # --- ADVANCED AI WEIGHTS PANEL ---
        with st.sidebar.expander("⚖️ Advanced AI Weights"):
            st.write("Adjust how strongly the AI penalizes bad groups or rewards good ones.")
            w = st.session_state.weights
            w['penalty_forced_location'] = st.number_input("Forced Location Broken (Penalty)", value=w['penalty_forced_location'], step=10000)
            w['penalty_must_go_with'] = st.number_input("'Must Go With' Broken (Penalty)", value=w['penalty_must_go_with'], step=10000)
            w['penalty_sole_gender'] = st.number_input("Sole Gender in Group (Penalty)", value=w['penalty_sole_gender'], step=10000)
            w['penalty_zero_friends'] = st.number_input("Zero Friends in Group (Penalty)", value=w['penalty_zero_friends'], step=10000)
            w['penalty_minority_gender_no_friends'] = st.number_input("Minority Gender + No Friends (Penalty)", value=w['penalty_minority_gender_no_friends'], step=10000)
            w['penalty_separation'] = st.number_input("'Can't Go With' Broken (Penalty)", value=w['penalty_separation'], step=10000)
            w['penalty_veto_activity'] = st.number_input("Activity Veto '1' Ignored (Penalty)", value=w['penalty_veto_activity'], step=10000)
            w['penalty_minority_gender_with_friends'] = st.number_input("Minority Gender + Has Friends (Penalty)", value=w['penalty_minority_gender_with_friends'], step=100)
            w['reward_activity_score_multiplier'] = st.number_input("Activity Score Multiplier (Preference Weight)", value=w['reward_activity_score_multiplier'], step=10)
            
            w['reward_friend_1'] = st.number_input("Reward for Keeping Friend 1", value=w.get('reward_friend_1', 120), step=10)
            w['reward_friend_2'] = st.number_input("Reward for Keeping Friend 2+", value=w.get('reward_friend_2', 100), step=10)
            
            w['penalty_group_size_diff_multiplier'] = st.number_input("Group Size Imbalance Multiplier", value=w['penalty_group_size_diff_multiplier'], step=1)
            
            if st.button("🔄 Reset Weights to Defaults"):
                st.session_state.weights = default_weights.copy()
                st.rerun()

        st.sidebar.markdown("---")

        # --- UI: NOT ATTENDING ---
        st.sidebar.header("🚫 Not Attending")
        not_attending_list = st.sidebar.multiselect("Select students:", options=all_students_list, key="y8_na_ms")
        st.session_state.not_attending = not_attending_list

        df_merged = df_merged_full[~df_merged_full['Official Name'].isin(not_attending_list)].copy()
        
        missing_students_df = df_stud[~df_stud['Email'].isin(df_resp['Email address'])].copy()
        missing_students_df = missing_students_df[~missing_students_df['Official Name'].isin(not_attending_list)]

        attending_students_list = sorted(df_merged['Official Name'].dropna().unique().tolist())

        # --- FIND STUDENT & REGENERATE BAR ---
        st.markdown("---")
        col_search, col_regen = st.columns([2, 1])
        with col_search:
            st.subheader("🔍 Find a Student")
            student_to_find = st.selectbox("Search for a student to quickly see their placement:", [""] + attending_students_list, key="find_student")
        with col_regen:
            st.subheader("⚙️ AI Controls")
            search_depth = st.select_slider("Search Depth (Time vs Perfection):", options=["Fast (1-2s)", "Standard (5-10s)", "Deep Search (20-40s)"], value="Standard (5-10s)")
            if st.button("🔄 Regenerate AI Groups", help="Forces the AI to calculate a completely new arrangement", use_container_width=True):
                st.session_state.generation_seed += 1
                st.rerun()

        find_result_container = st.empty()
        st.markdown("---")

        # --- UI: FORCE LOCATION ---
        st.sidebar.header("📍 Force Camp Location")
        force_stud = st.sidebar.selectbox("Student to Force", [""] + attending_students_list, key="f_stud")
        force_camp = st.sidebar.selectbox("Select Camp", ["", "Freycinet", "Bay of Fires"], key="f_camp")
        if st.sidebar.button("Add Location Rule"):
            if force_stud and force_camp:
                st.session_state.forced_locations[force_stud] = force_camp
                st.rerun() 
                
        if st.session_state.forced_locations:
            st.sidebar.write("**Active Location Rules:**")
            for s, c in list(st.session_state.forced_locations.items()): 
                col1, col2 = st.sidebar.columns([5, 1])
                col1.write(f"- {s} ➡️ {c}")
                if col2.button("❌", key=f"del_loc_{s}", help="Delete rule"):
                    del st.session_state.forced_locations[s]
                    st.rerun()
            if st.sidebar.button("Clear Location Rules"): st.session_state.forced_locations = {}; st.rerun()

        # --- UI: MUST GO WITH ---
        st.sidebar.header("🤝 'Must Go With' Rules")
        col3, col4 = st.sidebar.columns(2)
        with col3: must_1 = st.selectbox("Student A", [""] + attending_students_list, key="m1")
        with col4: must_2 = st.selectbox("Student B", [""] + attending_students_list, key="m2")
        if st.sidebar.button("Add 'Must Go' Rule"):
            if must_1 and must_2 and must_1 != must_2:
                st.session_state.must_go_pairs.append((must_1, must_2))
                st.rerun() 
                
        if st.session_state.must_go_pairs:
            st.sidebar.write("**Active 'Must Go' Rules:**")
            for i, pair in enumerate(st.session_state.must_go_pairs): 
                col1, col2 = st.sidebar.columns([5, 1])
                col1.write(f"- {pair[0]} & {pair[1]}")
                if col2.button("❌", key=f"del_mg_{i}", help="Delete rule"):
                    st.session_state.must_go_pairs.pop(i)
                    st.rerun()
            if st.sidebar.button("Clear 'Must Go' Rules"): st.session_state.must_go_pairs = []; st.rerun()

        # --- UI: CAN'T GO WITH ---
        st.sidebar.header("⚠️ 'Can't Go With' Rules")
        col5, col6 = st.sidebar.columns(2)
        with col5: sep_1 = st.selectbox("Student 1", [""] + attending_students_list, key="s1")
        with col6: sep_2 = st.selectbox("Student 2", [""] + attending_students_list, key="s2")
        if st.sidebar.button("Add 'Can't Go' Rule"):
            if sep_1 and sep_2 and sep_1 != sep_2:
                st.session_state.separation_pairs.append((sep_1, sep_2))
                st.rerun() 
                
        if st.session_state.separation_pairs:
            st.sidebar.write("**Active 'Can't Go' Rules:**")
            for i, pair in enumerate(st.session_state.separation_pairs): 
                col1, col2 = st.sidebar.columns([5, 1])
                col1.write(f"- {pair[0]} ≠ {pair[1]}")
                if col2.button("❌", key=f"del_sep_{i}", help="Delete rule"):
                    st.session_state.separation_pairs.pop(i)
                    st.rerun()
            if st.sidebar.button("Clear 'Can't Go' Rules"): st.session_state.separation_pairs = []; st.rerun()

        # --- SAVE SETTINGS FEATURE ---
        st.sidebar.markdown("---")
        st.sidebar.header("💾 Save Your Setup")
        
        current_settings = {
            "separation_pairs": st.session_state.separation_pairs,
            "must_go_pairs": st.session_state.must_go_pairs,
            "forced_locations": st.session_state.forced_locations,
            "not_attending": st.session_state.not_attending,
            "weights": st.session_state.weights 
        }
        st.sidebar.download_button(
            label="📥 Save Rules & Weights for Next Time (JSON)",
            data=json.dumps(current_settings, indent=4),
            file_name="camp_rules.json",
            mime="application/json"
        )

        # --- 3. CALCULATE ACTIVITY SCORES ---
        df_merged['Freycinet Score'] = pd.to_numeric(df_merged.get('How excited would you be to go sea kayaking?', 0), errors='coerce').fillna(0) + \
                                       pd.to_numeric(df_merged.get('How excited would you be to go coasteering?', 0), errors='coerce').fillna(0)
        df_merged['BoF Score'] = pd.to_numeric(df_merged.get('How excited would you be to go mountain bike riding on Blue (intermediate) flowy trails?', 0), errors='coerce').fillna(0) + \
                                 pd.to_numeric(df_merged.get('How excited would you be to go snorkelling?', 0), errors='coerce').fillna(0)
        
        def get_preference(row):
            if not row['Responded']: return 'No Response (Tie)'
            if row['Freycinet Score'] > row['BoF Score']: return 'Freycinet'
            elif row['BoF Score'] > row['Freycinet Score']: return 'Bay of Fires'
            else: return 'Tie'
        df_merged['Camp Preference'] = df_merged.apply(get_preference, axis=1)

        # --- 4. OPTIMIZATION ALGORITHM PER CLASS ---
        classes = sorted(df_merged['Which Connections class are you in? '].dropna().unique().tolist())
        styled_class_dfs = {}
        raw_class_dfs = {}
        class_friend_requests = {}
        leader_overview_data = []
        isolated_students = []

        if not_attending_list:
            st.warning(f"🚫 {len(not_attending_list)} student(s) marked as Not Attending. They have been removed from the sorting process.")

        st.header("Camp Group Assignments")

        # --- CROSS-CLASS RULE WARNING ---
        student_to_class = df_merged.set_index('Official Name')['Which Connections class are you in? '].to_dict()
        cross_class_warnings = []
        for s1, s2 in st.session_state.must_go_pairs + st.session_state.separation_pairs:
            c1 = student_to_class.get(s1)
            c2 = student_to_class.get(s2)
            if c1 and c2 and c1 != c2:
                cross_class_warnings.append(f"{s1} ({c1}) & {s2} ({c2})")
                
        if cross_class_warnings:
            st.warning("⚠️ **Warning: Rules across different classes detected!** The optimizer sorts class-by-class and cannot move a student to a different class. The following rules span multiple classes and will be ignored by the algorithm: " + ", ".join(cross_class_warnings))

        for cls in classes:
            class_df = df_merged[df_merged['Which Connections class are you in? '] == cls].copy()
            class_students = set(class_df['Official Name'])
            
            class_sep = [p for p in st.session_state.separation_pairs if p[0] in class_students or p[1] in class_students]
            class_must = [p for p in st.session_state.must_go_pairs if p[0] in class_students or p[1] in class_students]
            class_force = {k: v for k, v in st.session_state.forced_locations.items() if k in class_students}
            class_na = [s for s in st.session_state.not_attending if s in class_students]
            
            current_class_state = str({
                "sep": class_sep,
                "must": class_must,
                "force": class_force,
                "na": class_na,
                "w": st.session_state.weights,
                "seed": st.session_state.generation_seed,
                "depth": search_depth,
                "data_len": len(class_df)
            })

            if st.session_state.class_opt_states.get(cls) != current_class_state:
                if cls in st.session_state.optimized_groups:
                    del st.session_state.optimized_groups[cls] 
                st.session_state.class_opt_states[cls] = current_class_state

            G = nx.DiGraph()
            friend_requests = {}
            friend_cols = [c for c in class_df.columns if 'classmate' in c.lower()]
            
            for idx, row in class_df.iterrows():
                student = row['Official Name']
                
                sk = pd.to_numeric(row.get('How excited would you be to go sea kayaking?', 0), errors='coerce')
                coast = pd.to_numeric(row.get('How excited would you be to go coasteering?', 0), errors='coerce')
                mtb = pd.to_numeric(row.get('How excited would you be to go mountain bike riding on Blue (intermediate) flowy trails?', 0), errors='coerce')
                snork = pd.to_numeric(row.get('How excited would you be to go snorkelling?', 0), errors='coerce')
                
                f_veto = True if (sk == 1 or coast == 1) else False
                b_veto = True if (mtb == 1 or snork == 1) else False
                
                G.add_node(student, pref=row['Camp Preference'], f_score=row['Freycinet Score'], b_score=row['BoF Score'], gender=row.get('Gender', 'o'), f_veto=f_veto, b_veto=b_veto)
                
                reqs = []
                if row['Responded']:
                    for col in friend_cols:
                        friend_name = row.get(col)
                        if pd.notna(friend_name) and isinstance(friend_name, str) and friend_name.strip() != "" and friend_name.strip() not in not_attending_list:
                            reqs.append(friend_name.strip())
                friend_requests[student] = reqs
                
            for s1, s2 in st.session_state.must_go_pairs:
                if s1 in class_students and s2 in class_students:
                    if s1 in friend_requests and s2 not in friend_requests[s1]: friend_requests[s1].append(s2)
                    if s2 in friend_requests and s1 not in friend_requests[s2]: friend_requests[s2].append(s1)
                    
            for s1, s2 in st.session_state.separation_pairs:
                if s1 in friend_requests and s2 in friend_requests[s1]: friend_requests[s1].remove(s2)
                if s2 in friend_requests and s1 in friend_requests[s2]: friend_requests[s2].remove(s1)

            for student, reqs in friend_requests.items():
                valid_reqs = [f for f in reqs if f in class_students]
                friend_requests[student] = valid_reqs 
                for f_name in valid_reqs: 
                    G.add_edge(student, f_name)

            students = sorted(list(G.nodes))
            class_friend_requests[cls] = {k: list(v) for k, v in friend_requests.items()}

            if cls in st.session_state.optimized_groups:
                freycinet_group, bof_group = st.session_state.optimized_groups[cls]
            else:
                def calc_score(f_set, b_set):
                    w = st.session_state.weights 
                    score = 0
                    diff = abs(len(f_set) - len(b_set))
                    score -= (diff * diff * w['penalty_group_size_diff_multiplier']) 
                    if len(f_set) > 14 or len(b_set) > 14: score -= w['penalty_group_over_14']
                    
                    for f_stud, f_camp in st.session_state.forced_locations.items():
                        if f_camp == 'Freycinet' and f_stud in b_set: score -= w['penalty_forced_location']
                        elif f_camp == 'Bay of Fires' and f_stud in f_set: score -= w['penalty_forced_location']

                    for s1, s2 in st.session_state.must_go_pairs:
                        if (s1 in f_set and s2 in b_set) or (s1 in b_set and s2 in f_set): score -= w['penalty_must_go_with']

                    for s1, s2 in st.session_state.separation_pairs:
                        if (s1 in f_set and s2 in f_set) or (s1 in b_set and s2 in b_set): score -= w['penalty_separation']

                    for set_group in [f_set, b_set]:
                        m_count = sum(1 for s in set_group if G.nodes[s]['gender'] == 'm')
                        f_count = sum(1 for s in set_group if G.nodes[s]['gender'] == 'f')
                        for s in set_group:
                            is_f = set_group == f_set
                            
                            base_act_score = G.nodes[s]['f_score'] if is_f else G.nodes[s]['b_score']
                            score += (base_act_score * w['reward_activity_score_multiplier'])
                            
                            if is_f and G.nodes[s]['f_veto']: score -= w['penalty_veto_activity']
                            if not is_f and G.nodes[s]['b_veto']: score -= w['penalty_veto_activity']
                            
                            reqs = friend_requests.get(s, [])
                            friends_in_group = sum(1 for f in reqs if f in set_group) if reqs else 0
                            
                            if reqs:
                                if friends_in_group == 0: 
                                    score -= w['penalty_zero_friends'] 
                                else:
                                    for i, friend_name in enumerate(reqs):
                                        if friend_name in set_group:
                                            if i == 0: score += w.get('reward_friend_1', 120)
                                            else: score += w.get('reward_friend_2', 100)
                                
                            my_gen = G.nodes[s]['gender']
                            my_gen_count = m_count if my_gen == 'm' else (f_count if my_gen == 'f' else 0)
                            if my_gen in ['m', 'f']:
                                if my_gen_count == 1: score -= w['penalty_sole_gender'] 
                                elif 1 < my_gen_count < 4:
                                    if friends_in_group == 0: score -= w['penalty_minority_gender_no_friends'] 
                                    else: score -= w['penalty_minority_gender_with_friends'] 
                    return score

                best_overall_f = set()
                best_overall_b = set()
                best_overall_score = -float('inf')

                if "Fast" in search_depth: restarts, swaps = 50, 1000
                elif "Standard" in search_depth: restarts, swaps = 150, 2500
                else: restarts, swaps = 400, 3000

                with st.spinner(f"AI Optimizing Class {cls} ({restarts * swaps:,} checks)..."):
                    random.seed(f"camp_{cls}_fixed_{st.session_state.generation_seed}") 
                    
                    for restart in range(restarts):
                        shuffled_students = students.copy()
                        random.shuffle(shuffled_students)
                        curr_f = set(shuffled_students[:len(shuffled_students)//2])
                        curr_b = set(shuffled_students[len(shuffled_students)//2:])
                        curr_score = calc_score(curr_f, curr_b)
                        
                        for _ in range(swaps):
                            new_f, new_b = set(curr_f), set(curr_b)
                            if random.random() < 0.7 and len(new_f) > 0 and len(new_b) > 0:
                                s1 = random.choice(sorted(list(new_f)))
                                s2 = random.choice(sorted(list(new_b)))
                                new_f.remove(s1); new_f.add(s2)
                                new_b.remove(s2); new_b.add(s1)
                            else:
                                if random.random() < 0.5 and len(new_f) > 0:
                                    s = random.choice(sorted(list(new_f)))
                                    new_f.remove(s); new_b.add(s)
                                elif len(new_b) > 0:
                                    s = random.choice(sorted(list(new_b)))
                                    new_b.remove(s); new_f.add(s)
                            
                            new_score = calc_score(new_f, new_b)
                            if new_score > curr_score:
                                curr_score = new_score
                                curr_f, curr_b = new_f, new_b
                                
                        if curr_score > best_overall_score:
                            best_overall_score = curr_score
                            best_overall_f = curr_f
                            best_overall_b = curr_b

                freycinet_group = sorted(list(best_overall_f))
                bof_group = sorted(list(best_overall_b))
                st.session_state.optimized_groups[cls] = (freycinet_group, bof_group)

            # --- FIND STUDENT RESULT ---
            if student_to_find in freycinet_group:
                find_result_container.success(f"🎯 **{student_to_find}** was placed in **Freycinet** (Class {cls}). Scroll down to see their group!")
            elif student_to_find in bof_group:
                find_result_container.success(f"🎯 **{student_to_find}** was placed in **Bay of Fires** (Class {cls}). Scroll down to see their group!")

            # --- PREPARE DATAFRAME WITH VISUAL SPLITS ---
            class_roster = []
            
            for camp_name, camp_group in [('Freycinet', freycinet_group), ('Bay of Fires', bof_group)]:
                if camp_name == 'Bay of Fires':
                    class_roster.append({
                        'Student ID': '---', 'Email': '---', 'Student': f'--- BAY OF FIRES GROUP ---', 'Responded': '---', 'Gender': '---', 'Friend 1': '---', 'Friend 2': '---',
                        'Sea Kayak': '---', 'Coasteer': '---', 'Freycinet TOTAL': '---',
                        'MTB': '---', 'Snorkel': '---', 'Bay of Fires TOTAL': '---',
                        'Preferred Camp': '---', 'Assigned Camp': '---'
                    })
                    leader_overview_data.append({
                        'Class': f"--- {cls} ---", 'Assigned Camp': '--- BAY OF FIRES ---', 'Student ID': '---', 'Email': '---', 'Student': '---------------------------',
                        'Responded': '---', 'Sea Kayak': '---', 'Coasteer': '---', 'MTB Interest': '---', 'Snorkel': '---',
                        'General Camping Skill': '---', 'Sleeping Outdoors/Bugs': '---', 'Swimming Confidence': '---', 'Bike Comfort': '---', 'Overnight Hike': '---',
                        'Reaction to Hardship': '---', 'Gear Independence': '---', 'Group Teamwork': '---'
                    })
                else:
                    leader_overview_data.append({
                        'Class': f"--- {cls} ---", 'Assigned Camp': '--- FREYCINET ---', 'Student ID': '---', 'Email': '---', 'Student': '---------------------------',
                        'Responded': '---', 'Sea Kayak': '---', 'Coasteer': '---', 'MTB Interest': '---', 'Snorkel': '---',
                        'General Camping Skill': '---', 'Sleeping Outdoors/Bugs': '---', 'Swimming Confidence': '---', 'Bike Comfort': '---', 'Overnight Hike': '---',
                        'Reaction to Hardship': '---', 'Gear Independence': '---', 'Group Teamwork': '---'
                    })

                for student in camp_group:
                    student_row = class_df[class_df['Official Name'] == student].iloc[0]
                    reqs = friend_requests.get(student, [])
                    friend1 = reqs[0] if len(reqs) > 0 else ""
                    friend2 = reqs[1] if len(reqs) > 1 else ""
                    
                    friends_in_camp = sum(1 for f in reqs if f in camp_group)
                    if reqs and friends_in_camp == 0:
                        _why_parts = []
                        for _rf in reqs:
                            _other_grp = bof_group if camp_name == 'Freycinet' else freycinet_group
                            _is_sep = any(({student, _rf} == {a, b}) for a, b in st.session_state.separation_pairs)
                            if _is_sep:
                                _why_parts.append(f"{_rf}: separation rule applied")
                            elif _rf in st.session_state.forced_locations and st.session_state.forced_locations[_rf] != camp_name:
                                _why_parts.append(f"{_rf}: staff forced to {st.session_state.forced_locations[_rf]}")
                            elif camp_name == 'Freycinet' and G.nodes.get(_rf, {}).get('f_veto', False):
                                _why_parts.append(f"{_rf}: vetoed Freycinet activities")
                            elif camp_name == 'Bay of Fires' and G.nodes.get(_rf, {}).get('b_veto', False):
                                _why_parts.append(f"{_rf}: vetoed Bay of Fires activities")
                            elif _rf in _other_grp:
                                _why_parts.append(f"{_rf}: AI balanced group size / gender / competing requests")
                            else:
                                _why_parts.append(f"{_rf}: not placed in either group")
                        isolated_students.append({'Class': cls, 'Camp': camp_name, 'Student': student, 'Reason': 'Got 0 requested friends', 'Why': ' | '.join(_why_parts)})
                    elif not reqs:
                        isolated_students.append({'Class': cls, 'Camp': camp_name, 'Student': student, 'Reason': 'Made no friend requests', 'Why': 'No friend preferences submitted'})

                    sk = student_row.get('How excited would you be to go sea kayaking?', 'N/A') if student_row['Responded'] else 'N/A'
                    coast = student_row.get('How excited would you be to go coasteering?', 'N/A') if student_row['Responded'] else 'N/A'
                    mtb_int = student_row.get('How excited would you be to go mountain bike riding on Blue (intermediate) flowy trails?', 'N/A') if student_row['Responded'] else 'N/A'
                    snork = student_row.get('How excited would you be to go snorkelling?', 'N/A') if student_row['Responded'] else 'N/A'
                    mtb_ab = student_row.get(mtb_ability_col, 'N/A') if mtb_ability_col and student_row['Responded'] else 'N/A'
                    
                    class_roster.append({
                        'Student ID': student_row.get('Student ID', 'N/A'),
                        'Email': student_row.get('Email', 'N/A'),
                        'Student': student,
                        'Responded': 'Yes' if student_row['Responded'] else 'No',
                        'Gender': G.nodes[student]['gender'].upper(),
                        'Friend 1': friend1,
                        'Friend 2': friend2,
                        'Sea Kayak': sk,
                        'Coasteer': coast,
                        'Freycinet TOTAL': G.nodes[student]['f_score'],
                        'MTB': mtb_int,
                        'Snorkel': snork,
                        'Bay of Fires TOTAL': G.nodes[student]['b_score'],
                        'Preferred Camp': G.nodes[student]['pref'],
                        'Assigned Camp': camp_name
                    })
                    
                    leader_overview_data.append({
                        'Class': cls, 'Assigned Camp': camp_name, 
                        'Student ID': student_row.get('Student ID', 'N/A'),
                        'Email': student_row.get('Email', 'N/A'),
                        'Student': student,
                        'Responded': 'Yes' if student_row['Responded'] else 'No',
                        
                        'Sea Kayak': sk if camp_name == 'Freycinet' else '---',
                        'Coasteer': coast if camp_name == 'Freycinet' else '---',
                        'MTB Interest': mtb_int if camp_name == 'Bay of Fires' else '---',
                        'Snorkel': snork if camp_name == 'Bay of Fires' else '---',
                        
                        'General Camping Skill': student_row.get('How confident are you in your general camping skills (like setting up a tent or cooking on a camp stove)?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Sleeping Outdoors/Bugs': student_row.get('How do you feel about sleeping outdoors, using long-drop toilets, and being around dirt and bugs for a few days?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Swimming Confidence': student_row.get('How confident are you swimming in deep water where you cannot touch the bottom?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Bike Comfort': mtb_ab,
                        'Overnight Hike': student_row.get('How excited are you to go on an overnight hike carrying your gear?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Reaction to Hardship': student_row.get('When a physical activity gets tiring, difficult, or the weather turns bad, how do you usually react?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Gear Independence': student_row.get('How good are you at organising your own gear, packing your bag, and keeping yourself comfortable and safe without teachers or parents reminding you?', 'N/A') if student_row['Responded'] else 'N/A',
                        'Group Teamwork': student_row.get('How confident are you in helping your group work together, joining in discussions, and making classmates feel included?', 'N/A') if student_row['Responded'] else 'N/A'
                    })
            
            leader_overview_data.append({
                'Class': "", 'Assigned Camp': "", 'Student ID': "", 'Email': "", 'Student': "", 'Responded': "", 
                'Sea Kayak': "", 'Coasteer': "", 'MTB Interest': "", 'Snorkel': "",
                'General Camping Skill': "", 'Sleeping Outdoors/Bugs': "", 'Swimming Confidence': "", 'Bike Comfort': "", 'Overnight Hike': "",
                'Reaction to Hardship': "", 'Gear Independence': "", 'Group Teamwork': ""
            })
                
            df_class_final = pd.DataFrame(class_roster)

            # --- HIGHLIGHTING LOGIC ---
            def highlight_cells(row, df_reference, fr_dict):
                colors = [''] * len(row)
                
                if row['Student'] == student_to_find:
                    return ['background-color: #85e085; font-weight: bold; color: black; border: 2px solid green'] * len(row)

                if '---' in str(row['Student']):
                    return ['background-color: #d9d9d9; font-weight: bold; color: black; text-align: center'] * len(row)
                    
                student_to_camp = df_reference.set_index('Student')['Assigned Camp'].to_dict()
                
                if row['Responded'] == 'No': colors[df_reference.columns.get_loc('Responded')] = 'background-color: #ffffcc'
                
                for act_col in ['Sea Kayak', 'Coasteer', 'MTB', 'Snorkel']:
                    col_idx = df_reference.columns.get_loc(act_col)
                    val = row.get(act_col)
                    try:
                        if float(val) == 1.0:
                            if act_col in ['Sea Kayak', 'Coasteer'] and row['Assigned Camp'] == 'Freycinet':
                                colors[col_idx] = 'background-color: #ff4d4d; color: white; font-weight: bold'
                            elif act_col in ['MTB', 'Snorkel'] and row['Assigned Camp'] == 'Bay of Fires':
                                colors[col_idx] = 'background-color: #ff4d4d; color: white; font-weight: bold'
                    except:
                        pass

                camp_idx = df_reference.columns.get_loc('Assigned Camp')
                if row['Preferred Camp'] not in ['Tie', 'No Response (Tie)'] and row['Assigned Camp'] != row['Preferred Camp']:
                    colors[camp_idx] = 'background-color: #ffcccc' 
                    
                for f_col in ['Friend 1', 'Friend 2']:
                    f_idx = df_reference.columns.get_loc(f_col)
                    if row[f_col] != "":
                        f_camp = student_to_camp.get(row[f_col])
                        if f_camp is not None and f_camp != row['Assigned Camp']: colors[f_idx] = 'background-color: #ffcccc'
                            
                student_idx = df_reference.columns.get_loc('Student')
                
                reqs = fr_dict.get(row['Student'], [])
                if reqs:
                    friends_in_camp = sum(1 for f in reqs if student_to_camp.get(f) == row['Assigned Camp'])
                    if friends_in_camp == 0: colors[student_idx] = 'background-color: #ff9900; color: black; font-weight: bold'
                
                for s1, s2 in st.session_state.separation_pairs:
                    if row['Student'] == s1 and student_to_camp.get(s2) == row['Assigned Camp']: colors[student_idx] = 'background-color: #ff4d4d'
                    elif row['Student'] == s2 and student_to_camp.get(s1) == row['Assigned Camp']: colors[student_idx] = 'background-color: #ff4d4d'

                my_gender = row['Gender']
                if my_gender in ['M', 'F']:
                    gender_count_in_camp = sum((df_reference['Assigned Camp'] == row['Assigned Camp']) & (df_reference['Gender'] == my_gender))
                    gen_idx = df_reference.columns.get_loc('Gender')
                    if gender_count_in_camp == 1:
                        colors[gen_idx] = 'background-color: #ff4d4d; color: white' 
                    elif 1 < gender_count_in_camp < 4:
                        colors[gen_idx] = 'background-color: #ffcc00; color: black' 

                return colors

            styled_df = df_class_final.style.apply(highlight_cells, df_reference=df_class_final, fr_dict=friend_requests, axis=1)
            styled_class_dfs[cls] = styled_df
            raw_class_dfs[cls] = df_class_final.copy()

            st.subheader(f"Class: {cls}")
            f_males = sum(1 for s in freycinet_group if G.nodes[s]['gender'] == 'm')
            f_females = sum(1 for s in freycinet_group if G.nodes[s]['gender'] == 'f')
            b_males = sum(1 for s in bof_group if G.nodes[s]['gender'] == 'm')
            b_females = sum(1 for s in bof_group if G.nodes[s]['gender'] == 'f')
            
            st.write(f"**Freycinet:** {len(freycinet_group)} students ({f_males}M, {f_females}F) | **Bay of Fires:** {len(bof_group)} students ({b_males}M, {b_females}F)")
            st.dataframe(styled_df, use_container_width=True)

        # --- 5. DASHBOARDS ---
        colA, colB = st.columns(2)
        with colA:
            st.write("### ⚠️ At Risk of Isolation")
            if isolated_students: st.dataframe(pd.DataFrame(isolated_students), use_container_width=True)
            else: st.success("No isolated students!")
                
        with colB:
            st.write("### ❌ Missing Responses List")
            if not missing_students_df.empty: st.dataframe(missing_students_df[['First name', 'Surname', 'Rollgroup', 'Email']], use_container_width=True)
            else: st.success("Everyone has responded!")

        # --- LEADER OVERVIEW STYLING ---
        df_leaders = pd.DataFrame(leader_overview_data)
        
        def highlight_leader_overview(row, df_ref):
            colors = [''] * len(row)
            
            if '---' in str(row.get('Student', '')):
                return ['background-color: #d9d9d9; font-weight: bold; text-align: center'] * len(row)
            
            cols_to_check = [
                'General Camping Skill', 'Sleeping Outdoors/Bugs', 'Swimming Confidence', 'Bike Comfort', 'Overnight Hike', 
                'Reaction to Hardship', 'Gear Independence', 'Group Teamwork',
                'Sea Kayak', 'Coasteer', 'MTB Interest', 'Snorkel'
            ]

            for col_name in cols_to_check:
                if col_name in df_ref.columns:
                    col_idx = df_ref.columns.get_loc(col_name)
                    val = str(row[col_name]).strip()
                    match = re.search(r'^\s*([1-5])(?:\D|$)', val)
                    if match:
                        num = int(match.group(1))
                        if num in [1, 2]: colors[col_idx] = 'background-color: #a83232; color: white; font-weight: bold' # Dark Red
                        elif num in [3, 4]: colors[col_idx] = 'background-color: #ff9933; color: black; font-weight: bold' # Orange
                        elif num == 5: colors[col_idx] = 'background-color: #ffcc00; color: black; font-weight: bold' # Yellow
            return colors

        styled_leaders = df_leaders.style.apply(highlight_leader_overview, df_ref=df_leaders, axis=1)

        # ─── Excel colour helpers ─────────────────────────────────────────────────────
        def _apply_y8_excel_colors(ws, df_data, fr_map, sep_pairs, find_name):
            """Apply highlight colours matching the on-screen display to a Y8 class sheet."""
            fill_grey       = PatternFill("solid", fgColor="D9D9D9")
            fill_green      = PatternFill("solid", fgColor="85E085")
            fill_yellow     = PatternFill("solid", fgColor="FFFFCC")
            fill_red        = PatternFill("solid", fgColor="FF4D4D")
            fill_pink       = PatternFill("solid", fgColor="FFCCCC")
            fill_orange     = PatternFill("solid", fgColor="FF9900")
            fill_amber      = PatternFill("solid", fgColor="FFCC00")
            font_bold       = Font(bold=True)
            font_bold_white = Font(bold=True, color="FFFFFF")
            # Style header row
            for col_idx in range(1, len(df_data.columns) + 1):
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="4472C4")
                ws.cell(row=1, column=col_idx).font = Font(bold=True, color="FFFFFF")
            cols = list(df_data.columns)
            # Build student→camp lookup (skip separator rows)
            student_to_camp_xl = {}
            for _, row in df_data.iterrows():
                if '---' not in str(row.get('Student', '')):
                    student_to_camp_xl[row['Student']] = str(row.get('Assigned Camp', ''))
            non_sep = df_data[~df_data['Student'].astype(str).str.contains('---', na=False)]
            for excel_row, (_, row) in enumerate(df_data.iterrows(), start=2):
                student  = str(row.get('Student', ''))
                assigned = str(row.get('Assigned Camp', ''))
                # Separator rows — full grey
                if '---' in student:
                    for c in range(1, len(cols) + 1):
                        ws.cell(row=excel_row, column=c).fill = fill_grey
                        ws.cell(row=excel_row, column=c).font = font_bold
                    continue
                # Searched-for student — full green, highest priority
                if find_name and student == find_name:
                    for c in range(1, len(cols) + 1):
                        ws.cell(row=excel_row, column=c).fill = fill_green
                        ws.cell(row=excel_row, column=c).font = font_bold
                    continue
                # Responded column — yellow if no response
                if row.get('Responded') == 'No' and 'Responded' in cols:
                    ws.cell(row=excel_row, column=cols.index('Responded') + 1).fill = fill_yellow
                # Activity veto (score 1 in a forced activity) — red
                for act_col, veto_camp in [('Sea Kayak', 'Freycinet'), ('Coasteer', 'Freycinet'),
                                            ('MTB', 'Bay of Fires'), ('Snorkel', 'Bay of Fires')]:
                    if act_col in cols:
                        try:
                            if float(row[act_col]) == 1.0 and assigned == veto_camp:
                                c = ws.cell(row=excel_row, column=cols.index(act_col) + 1)
                                c.fill = fill_red; c.font = font_bold_white
                        except (ValueError, TypeError):
                            pass
                # Wrong camp preference — pink on Assigned Camp
                if 'Preferred Camp' in cols and 'Assigned Camp' in cols:
                    pref = str(row.get('Preferred Camp', ''))
                    if pref not in ('Tie', 'No Response (Tie)') and pref and assigned != pref:
                        ws.cell(row=excel_row, column=cols.index('Assigned Camp') + 1).fill = fill_pink
                # Friend not in same camp — pink on friend column
                for f_col in ('Friend 1', 'Friend 2'):
                    if f_col in cols:
                        f_name_val = str(row.get(f_col, ''))
                        if f_name_val not in ('', 'nan', '---'):
                            f_camp = student_to_camp_xl.get(f_name_val)
                            if f_camp is not None and f_camp != assigned:
                                ws.cell(row=excel_row, column=cols.index(f_col) + 1).fill = fill_pink
                # Student column: isolation (0 friends) = orange; separation violation = red
                if 'Student' in cols:
                    reqs = fr_map.get(student, [])
                    if reqs:
                        friends_in_camp = sum(1 for f in reqs if student_to_camp_xl.get(f) == assigned)
                        if friends_in_camp == 0:
                            c = ws.cell(row=excel_row, column=cols.index('Student') + 1)
                            c.fill = fill_orange; c.font = font_bold
                    for s1, s2 in sep_pairs:
                        if student == s1 and student_to_camp_xl.get(s2) == assigned:
                            ws.cell(row=excel_row, column=cols.index('Student') + 1).fill = fill_red
                        elif student == s2 and student_to_camp_xl.get(s1) == assigned:
                            ws.cell(row=excel_row, column=cols.index('Student') + 1).fill = fill_red
                # Gender balance — red if sole gender, amber if 2-3
                if 'Gender' in cols:
                    my_gen = str(row.get('Gender', ''))
                    if my_gen in ('M', 'F'):
                        g_count = int(((non_sep['Assigned Camp'] == assigned) & (non_sep['Gender'] == my_gen)).sum())
                        gen_cell = ws.cell(row=excel_row, column=cols.index('Gender') + 1)
                        if g_count == 1:
                            gen_cell.fill = fill_red; gen_cell.font = font_bold_white
                        elif 1 < g_count < 4:
                            gen_cell.fill = fill_amber

        def _apply_leader_excel_colors(ws, df_data):
            """Apply skill-score colours to the Leader Overview sheet."""
            fill_grey       = PatternFill("solid", fgColor="D9D9D9")
            fill_dark_red   = PatternFill("solid", fgColor="A83232")
            fill_orange     = PatternFill("solid", fgColor="FF9933")
            fill_yellow     = PatternFill("solid", fgColor="FFCC00")
            font_bold_white = Font(bold=True, color="FFFFFF")
            font_bold       = Font(bold=True)
            for col_idx in range(1, len(df_data.columns) + 1):
                ws.cell(row=1, column=col_idx).fill = PatternFill("solid", fgColor="4472C4")
                ws.cell(row=1, column=col_idx).font = Font(bold=True, color="FFFFFF")
            cols = list(df_data.columns)
            skill_cols = ['General Camping Skill', 'Sleeping Outdoors/Bugs', 'Swimming Confidence',
                          'Bike Comfort', 'Overnight Hike', 'Reaction to Hardship',
                          'Gear Independence', 'Group Teamwork', 'Sea Kayak', 'Coasteer',
                          'MTB Interest', 'Snorkel']
            for excel_row, (_, row) in enumerate(df_data.iterrows(), start=2):
                if '---' in str(row.get('Student', '')):
                    for c in range(1, len(cols) + 1):
                        ws.cell(row=excel_row, column=c).fill = fill_grey
                        ws.cell(row=excel_row, column=c).font = font_bold
                    continue
                for col_name in skill_cols:
                    if col_name in cols:
                        val = str(row.get(col_name, '')).strip()
                        m = re.search(r'^\s*([1-5])(?:\D|$)', val)
                        if m:
                            num = int(m.group(1))
                            cell = ws.cell(row=excel_row, column=cols.index(col_name) + 1)
                            if num in (1, 2):
                                cell.fill = fill_dark_red; cell.font = font_bold_white
                            elif num in (3, 4):
                                cell.fill = fill_orange; cell.font = font_bold
                            elif num == 5:
                                cell.fill = fill_yellow; cell.font = font_bold

        # --- 6. EXCEL EXPORT ---
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for cls, df_raw in raw_class_dfs.items():
                sheet_name = str(cls)[:31]
                df_raw.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                _apply_y8_excel_colors(ws, df_raw, class_friend_requests.get(cls, {}),
                                       st.session_state.separation_pairs, student_to_find)
                for col in ws.columns:
                    ws.column_dimensions[col[0].column_letter].width = max(
                        (len(str(cell.value or '')) for cell in col), default=0) + 2

            # LEADER OVERVIEW
            df_leaders.to_excel(writer, sheet_name="Leader Overview", index=False)
            ws_leaders = writer.sheets["Leader Overview"]
            _apply_leader_excel_colors(ws_leaders, df_leaders)
            for col in ws_leaders.columns:
                ws_leaders.column_dimensions[col[0].column_letter].width = min(
                    max((len(str(cell.value or '')) for cell in col), default=0) + 2, 50)

            # Add Not Attending Tab
            if not_attending_list:
                na_data = df_merged_full[df_merged_full['Official Name'].isin(not_attending_list)]
                na_export_df = na_data[['Student ID', 'Email', 'Official Name', 'Which Connections class are you in? ']].copy()
                na_export_df.rename(columns={
                    'Official Name': 'Student', 
                    'Which Connections class are you in? ': 'Class Group'
                }, inplace=True)
                
                na_export_df.to_excel(writer, sheet_name="Not Attending", index=False)
                worksheet_na = writer.sheets["Not Attending"]
                for col in worksheet_na.columns:
                    worksheet_na.column_dimensions[col[0].column_letter].width = max((len(str(cell.value)) for cell in col), default=0) + 2

        output.seek(0)
        st.markdown("---")
        st.download_button(label="📥 Download Final Excel File", data=output, file_name="Camp_Allocations_Final.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    else:
        st.info("👈 Please upload both CSV files in the sidebar to begin.")


# =========================================================================================
# ========================= PAGE 2: FINAL PUBLIC ROSTER BUILDER ===========================
# =========================================================================================
elif page == "📋 Final Roster & Leader Builder":
    st.title("📋 Final Roster & Leader Builder")

    roster_tab_y8, roster_tab_y9 = st.tabs(["🏕️ Year 8 Camps", "🏔️ Year 9 Camps"])

    # ── Year 8 Roster ─────────────────────────────────────────────────────────────────────
    with roster_tab_y8:
        st.subheader("Year 8 — Final Public Roster & Leader Builder")
        st.info("Upload your manually tweaked `Camp_Allocations_Final.xlsx` file here. The app will visually scan where you dragged the students in Excel, build a privacy-safe public roster, and completely regenerate an up-to-date Leader Overview for your staff.")

        if df_merged_full is None:
            st.warning("⚠️ Please upload the Y8 Form Responses and Y8 Student List CSVs in the sidebar (Step 1) to use this tool.")
        else:
            uploaded_xlsx = st.file_uploader("Upload Modified Excel File (.xlsx)", type=["xlsx"])
        
            if uploaded_xlsx:
                xls = pd.ExcelFile(uploaded_xlsx)
                sheet_names = xls.sheet_names
            
                master_assignments = []
                not_attending_ids = []
            
                # Create a set of all genuinely known Student IDs from the master list (to filter out random teacher notes!)
                valid_student_ids = set(df_merged_full['Student ID'].dropna().astype(str).str.replace(r'\.0$', '', regex=True).str.strip().tolist())
                valid_student_ids.discard('nan')
                valid_student_ids.discard('N/A')
                valid_student_ids.discard('')
            
                # 1. Capture the Not Attending list directly from the Excel file
                if "Not Attending" in sheet_names:
                    df_na = pd.read_excel(xls, sheet_name="Not Attending")
                    if 'Student ID' in df_na.columns:
                        not_attending_ids = df_na['Student ID'].dropna().astype(str).str.replace(r'\.0$', '', regex=True).tolist()
            
                # 2. Visually scan the Class tabs to see where the user placed the students
                for sheet in sheet_names:
                    if sheet in ["Leader Overview", "Not Attending"]:
                        continue # Skip non-class sheets
                
                    df_class = pd.read_excel(xls, sheet_name=sheet)
                    if 'Student ID' in df_class.columns:
                        current_camp = "Freycinet" # Default top section
                    
                        for _, row in df_class.iterrows():
                            stud_val = str(row.get('Student', '')).strip()
                            st_id = str(row.get('Student ID', '')).replace('.0', '').strip()
                        
                            # When the scanner hits the visual separator, flip the active camp
                            if 'BAY OF FIRES' in stud_val.upper() or 'BAY OF FIRES' in st_id.upper():
                                current_camp = "Bay of Fires"
                                continue
                            
                            # Save the assignment - ONLY if it's a verified valid Student ID (ignores ALL teacher notes/blanks perfectly)
                            if st_id in valid_student_ids and st_id not in not_attending_ids:
                                master_assignments.append({
                                    'Class': str(sheet).strip(),
                                    'Student ID': st_id,
                                    'Assigned Camp': current_camp
                                })
            
                if not master_assignments:
                    st.error("Could not find any valid student assignments in the uploaded Excel file.")
                else:
                    df_assignments = pd.DataFrame(master_assignments)
                    st.success(f"✅ Successfully read data for {len(df_assignments)} students across {df_assignments['Class'].nunique()} classes!")
                
                    # --- A. BUILD PUBLIC ROSTER (DATAFRAME FOR PREVIEW) ---
                    df_pub = pd.merge(df_assignments, df_stud_pub[['Student ID', 'Preferred name', 'Surname']], on='Student ID', how='left')
                    df_pub = df_pub[['Class', 'Preferred name', 'Surname', 'Assigned Camp']]
                    df_pub['Full Name'] = df_pub['Preferred name'].astype(str) + " " + df_pub['Surname'].astype(str)
                
                    # --- NEW FEATURE: GENERATE VISUALLY APPEALING PUBLIC EXCEL ROSTER ---
                    def generate_public_excel(df_roster):
                        wb = Workbook()
                        ws = wb.active
                        ws.title = "Camp Roster"

                        # Styles
                        title_font = Font(size=16, bold=True, color="FFFFFF")
                        f_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
                        b_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid") # Red
                        date_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid") # Dark Grey
                        class_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Light Grey
                    
                        center_align = Alignment(horizontal="center", vertical="center")
                        left_align = Alignment(horizontal="left", vertical="center")
                        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                        unique_classes = sorted(df_roster['Class'].unique().tolist())
                    
                        # Dynamically match classes based on user's exact keywords
                        g1_names = ['rustin', 'stowe', 'preston']
                        g2_names = ['bracey', 'knight', 'tuke']
                    
                        g1_actual = [c for c in unique_classes if any(k in c.lower() for k in g1_names)]
                        g2_actual = [c for c in unique_classes if any(k in c.lower() for k in g2_names)]
                    
                        # Fallback if class names don't match those keywords at all
                        if not g1_actual and not g2_actual:
                            mid = len(unique_classes) // 2
                            g1_actual = unique_classes[:mid]
                            g2_actual = unique_classes[mid:]

                        def write_block(ws, start_row, date_text, classes):
                            if not classes: return start_row
                        
                            f_end_col = len(classes)
                            total_cols = f_end_col * 2
                        
                            # Row 1: Date Header
                            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
                            cell = ws.cell(row=start_row, column=1, value=date_text)
                            cell.font = title_font
                            cell.fill = date_fill
                            cell.alignment = center_align
                        
                            # Row 2: Camp Locations
                            ws.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=f_end_col)
                            cell_f = ws.cell(row=start_row+1, column=1, value="FREYCINET")
                            cell_f.font = Font(size=14, bold=True, color="FFFFFF")
                            cell_f.fill = f_fill
                            cell_f.alignment = center_align
                        
                            ws.merge_cells(start_row=start_row+1, start_column=f_end_col+1, end_row=start_row+1, end_column=total_cols)
                            cell_b = ws.cell(row=start_row+1, column=f_end_col+1, value="BAY OF FIRES")
                            cell_b.font = Font(size=14, bold=True, color="FFFFFF")
                            cell_b.fill = b_fill
                            cell_b.alignment = center_align
                        
                            # Row 3: Class Headers
                            for i, cls_name in enumerate(classes):
                                # Freycinet Side
                                c1 = ws.cell(row=start_row+2, column=i+1, value=cls_name)
                                c1.font = Font(bold=True)
                                c1.fill = class_fill
                                c1.alignment = center_align
                                c1.border = thin_border
                            
                                # Bay of Fires Side
                                c2 = ws.cell(row=start_row+2, column=f_end_col+i+1, value=cls_name)
                                c2.font = Font(bold=True)
                                c2.fill = class_fill
                                c2.alignment = center_align
                                c2.border = thin_border
                        
                            # Row 4+: Student Names
                            f_lists = []
                            b_lists = []
                            for cls_name in classes:
                                f_students = df_roster[(df_roster['Class'] == cls_name) & (df_roster['Assigned Camp'] == 'Freycinet')]['Full Name'].tolist()
                                b_students = df_roster[(df_roster['Class'] == cls_name) & (df_roster['Assigned Camp'] == 'Bay of Fires')]['Full Name'].tolist()
                                f_lists.append(f_students)
                                b_lists.append(b_students)
                        
                            max_len = max([len(l) for l in f_lists + b_lists] + [0])
                        
                            for row_offset in range(max_len):
                                for col_idx, f_list in enumerate(f_lists):
                                    val = f_list[row_offset] if row_offset < len(f_list) else ""
                                    cell = ws.cell(row=start_row+3+row_offset, column=col_idx+1, value=val)
                                    cell.alignment = left_align
                                    cell.border = thin_border
                                
                                for col_idx, b_list in enumerate(b_lists):
                                    val = b_list[row_offset] if row_offset < len(b_list) else ""
                                    cell = ws.cell(row=start_row+3+row_offset, column=f_end_col+col_idx+1, value=val)
                                    cell.alignment = left_align
                                    cell.border = thin_border
                                
                            # Return the next available row (with a 2-row buffer space)
                            return start_row + 3 + max_len + 2
                        
                        current_row = 2
                        if g1_actual:
                            current_row = write_block(ws, current_row, "Camp Dates: 11 - 15 May", g1_actual)
                        if g2_actual:
                            current_row = write_block(ws, current_row, "Camp Dates: 18 - 22 May", g2_actual)
                        
                        # Handle any extra stray classes just in case
                        other_actual = [c for c in unique_classes if c not in g1_actual and c not in g2_actual]
                        if other_actual:
                            for i in range(0, len(other_actual), 3):
                                chunk = other_actual[i:i+3]
                                current_row = write_block(ws, current_row, "Other Classes", chunk)
                            
                        # Set standard column widths
                        max_cols = (max(len(g1_actual), len(g2_actual), 1) * 2)
                        for col in range(1, max_cols + 1):
                            ws.column_dimensions[get_column_letter(col)].width = 25
                        
                        return wb

                    wb_public = generate_public_excel(df_pub)
                    output_pub = io.BytesIO()
                    wb_public.save(output_pub)
                    output_pub.seek(0)
                
                    # --- B. BUILD UPDATED LEADER OVERVIEW ---
                    leader_overview_data = []
                
                    for cls, class_group in df_assignments.groupby('Class'):
                        for camp_name in ['Freycinet', 'Bay of Fires']:
                            camp_students = class_group[class_group['Assigned Camp'] == camp_name]['Student ID'].tolist()
                            if not camp_students: continue
                        
                            # Add separator
                            leader_overview_data.append({
                                'Class': f"--- {cls} ---", 'Assigned Camp': f'--- {camp_name.upper()} ---', 'Student ID': '---', 'Email': '---', 'Student': '---------------------------',
                                'Responded': '---', 'Sea Kayak': '---', 'Coasteer': '---', 'MTB Interest': '---', 'Snorkel': '---',
                                'General Camping Skill': '---', 'Sleeping Outdoors/Bugs': '---', 'Swimming Confidence': '---', 'Bike Comfort': '---', 'Overnight Hike': '---',
                                'Reaction to Hardship': '---', 'Gear Independence': '---', 'Group Teamwork': '---'
                            })
                        
                            for st_id in camp_students:
                                stud_rows = df_merged_full[df_merged_full['Student ID'] == st_id]
                                if stud_rows.empty: continue
                                student_row = stud_rows.iloc[0]
                            
                                sk = student_row.get('How excited would you be to go sea kayaking?', 'N/A') if student_row['Responded'] else 'N/A'
                                coast = student_row.get('How excited would you be to go coasteering?', 'N/A') if student_row['Responded'] else 'N/A'
                                mtb_int = student_row.get('How excited would you be to go mountain bike riding on Blue (intermediate) flowy trails?', 'N/A') if student_row['Responded'] else 'N/A'
                                snork = student_row.get('How excited would you be to go snorkelling?', 'N/A') if student_row['Responded'] else 'N/A'
                                mtb_ab = student_row.get(mtb_ability_col, 'N/A') if mtb_ability_col and student_row['Responded'] else 'N/A'
                            
                                leader_overview_data.append({
                                    'Class': cls, 'Assigned Camp': camp_name, 
                                    'Student ID': st_id,
                                    'Email': student_row.get('Email', 'N/A'),
                                    'Student': student_row['Official Name'],
                                    'Responded': 'Yes' if student_row['Responded'] else 'No',
                                
                                    'Sea Kayak': sk if camp_name == 'Freycinet' else '---',
                                    'Coasteer': coast if camp_name == 'Freycinet' else '---',
                                    'MTB Interest': mtb_int if camp_name == 'Bay of Fires' else '---',
                                    'Snorkel': snork if camp_name == 'Bay of Fires' else '---',
                                
                                    'General Camping Skill': student_row.get('How confident are you in your general camping skills (like setting up a tent or cooking on a camp stove)?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Sleeping Outdoors/Bugs': student_row.get('How do you feel about sleeping outdoors, using long-drop toilets, and being around dirt and bugs for a few days?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Swimming Confidence': student_row.get('How confident are you swimming in deep water where you cannot touch the bottom?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Bike Comfort': mtb_ab,
                                    'Overnight Hike': student_row.get('How excited are you to go on an overnight hike carrying your gear?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Reaction to Hardship': student_row.get('When a physical activity gets tiring, difficult, or the weather turns bad, how do you usually react?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Gear Independence': student_row.get('How good are you at organising your own gear, packing your bag, and keeping yourself comfortable and safe without teachers or parents reminding you?', 'N/A') if student_row['Responded'] else 'N/A',
                                    'Group Teamwork': student_row.get('How confident are you in helping your group work together, joining in discussions, and making classmates feel included?', 'N/A') if student_row['Responded'] else 'N/A'
                                })
                            
                        leader_overview_data.append({
                            'Class': "", 'Assigned Camp': "", 'Student ID': "", 'Email': "", 'Student': "", 'Responded': "", 
                            'Sea Kayak': "", 'Coasteer': "", 'MTB Interest': "", 'Snorkel': "",
                            'General Camping Skill': "", 'Sleeping Outdoors/Bugs': "", 'Swimming Confidence': "", 'Bike Comfort': "", 'Overnight Hike': "",
                            'Reaction to Hardship': "", 'Gear Independence': "", 'Group Teamwork': ""
                        })
                    
                    df_leader_updated = pd.DataFrame(leader_overview_data)
                
                    # --- C. DISPLAY AND EXPORT ---
                    col1, col2 = st.columns(2)
                
                    with col1:
                        st.subheader("📋 Public Roster Processing Complete")
                        st.write("The simple flat list is shown below, but the exported Excel file is styled horizontally into groups (May 11-15 and May 18-22) for easy reading!")
                        st.dataframe(df_pub[['Class', 'Full Name', 'Assigned Camp']], hide_index=True, use_container_width=True)
                    
                        st.download_button(
                            label="📥 Download Stylized Public Roster (Excel)", 
                            data=output_pub, 
                            file_name="Final_Public_Camp_Roster.xlsx", 
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", 
                            use_container_width=True
                        )
                    
                    with col2:
                        st.subheader("🛡️ Updated Leader Overview Preview")
                    
                        def highlight_leader_overview(row, df_ref):
                            colors = [''] * len(row)
                            if '---' in str(row.get('Student', '')):
                                return ['background-color: #d9d9d9; font-weight: bold; text-align: center'] * len(row)
                        
                            cols_to_check = ['General Camping Skill', 'Sleeping Outdoors/Bugs', 'Swimming Confidence', 'Bike Comfort', 'Overnight Hike', 'Reaction to Hardship', 'Gear Independence', 'Group Teamwork', 'Sea Kayak', 'Coasteer', 'MTB Interest', 'Snorkel']
                            for col_name in cols_to_check:
                                if col_name in df_ref.columns:
                                    col_idx = df_ref.columns.get_loc(col_name)
                                    val = str(row[col_name]).strip()
                                    match = re.search(r'^\s*([1-5])(?:\D|$)', val)
                                    if match:
                                        num = int(match.group(1))
                                        if num in [1, 2]: colors[col_idx] = 'background-color: #a83232; color: white; font-weight: bold'
                                        elif num in [3, 4]: colors[col_idx] = 'background-color: #ff9933; color: black; font-weight: bold'
                                        elif num == 5: colors[col_idx] = 'background-color: #ffcc00; color: black; font-weight: bold'
                            return colors

                        styled_leaders = df_leader_updated.style.apply(highlight_leader_overview, df_ref=df_leader_updated, axis=1)
                        st.dataframe(styled_leaders, use_container_width=True)
                    
                        output_ldr = io.BytesIO()
                        with pd.ExcelWriter(output_ldr, engine='openpyxl') as writer:
                            styled_leaders.to_excel(writer, sheet_name="Updated Leader Overview", index=False)
                            worksheet_l = writer.sheets["Updated Leader Overview"]
                            for col in worksheet_l.columns: worksheet_l.column_dimensions[col[0].column_letter].width = min(max((len(str(cell.value)) for cell in col), default=0) + 2, 50)
                        
                            if not_attending_ids:
                                na_data = df_merged_full[df_merged_full['Student ID'].isin(not_attending_ids)]
                                na_export_df = na_data[['Student ID', 'Email', 'Official Name', 'Which Connections class are you in? ']].copy()
                                na_export_df.rename(columns={'Official Name': 'Student', 'Which Connections class are you in? ': 'Class Group'}, inplace=True)
                                na_export_df.to_excel(writer, sheet_name="Not Attending", index=False)
                                worksheet_na = writer.sheets["Not Attending"]
                                for col in worksheet_na.columns: worksheet_na.column_dimensions[col[0].column_letter].width = max((len(str(cell.value)) for cell in col), default=0) + 2
                            
                        output_ldr.seek(0)
                        st.download_button(label="📥 Download Updated Leader Overview (Excel)", data=output_ldr, file_name="Updated_Leader_Overview.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True)

    # ── Year 9 Roster ─────────────────────────────────────────────────────────────────────
    with roster_tab_y9:
        st.subheader("Year 9 — Final Public Roster Builder")
        st.info("Upload the `Y9_Journey_Groups.xlsx` file exported from the **Y9 Journey Groups** tool. The app will read each camp group sheet and build a clean public-facing roster organised by week and camp.")

        y9_roster_xlsx = st.file_uploader(
            "Upload your edited Y9_Journey_Groups.xlsx", type=["xlsx"], key="y9_roster_upload",
            help="Export this file from the Y9 Journey Groups tool, adjust groups in Excel, then upload it here to build the public roster.")

        if y9_roster_xlsx:
            # ── Camp key lookup ────────────────────────────────────────────────────────────
            _Y9R_CAMP_LABELS = {
                'MTB':        'Mountain to Sea MTB',
                'CC':         'Colossal Cliffs',
                'MM':         'Mersey & Mountains',
                'Explorer':   'Cradle to Coast Explorer',
                'Challenger': 'Cradle to Coast Challenger',
            }
            _Y9R_COLORS = {
                'MTB':        'FF2E7D32',
                'CC':         'FF1565C0',
                'MM':         'FF6A1B9A',
                'Explorer':   'FFE65100',
                'Challenger': 'FFB71C1C',
            }
            _WEEK_LABELS = {
                '1': 'Week 1: Unwin & Hodgkin (12–18 November)',
                '2': 'Week 2: Mather & Ransome (23–29 November)',
            }

            _xls_y9r = pd.ExcelFile(y9_roster_xlsx)
            _sheets_y9r = [s for s in _xls_y9r.sheet_names
                           if s not in ('Isolation Report', 'Not Attending')]

            # Parse sheets: name format  CAMPKEY[_A|_B]_W1 or CAMPKEY_W1
            _all_assignments = []
            for _sheet in _sheets_y9r:
                _parts = _sheet.split('_')
                _camp_key = _parts[0] if _parts else ''
                _week_num = _parts[-1].replace('W', '') if _parts else ''
                _subgroup = _parts[1] if len(_parts) == 3 else ''

                _df_sh = pd.read_excel(_xls_y9r, sheet_name=_sheet)
                for _, _row in _df_sh.iterrows():
                    _name = str(_row.get('Student', '')).strip()
                    if _name and _name.lower() not in ('nan', ''):
                        _all_assignments.append({
                            'Student':   _name,
                            'Camp Key':  _camp_key,
                            'Camp':      _Y9R_CAMP_LABELS.get(_camp_key, _camp_key),
                            'Group':     _subgroup if _subgroup else 'A',
                            'Week':      _week_num,
                            'Week Label': _WEEK_LABELS.get(_week_num, f'Week {_week_num}'),
                        })

            if not _all_assignments:
                st.error("Could not read any student assignments from the uploaded file. Check that it was exported from the Y9 Journey Groups tool.")
            else:
                _df_asgn = pd.DataFrame(_all_assignments)
                st.success(f"✅ Read {len(_df_asgn)} student assignments across {_df_asgn['Camp'].nunique()} camps and {_df_asgn['Week'].nunique()} weeks.")

                # ── Preview table ──────────────────────────────────────────────────────────
                for _wk in sorted(_df_asgn['Week'].unique()):
                    _wk_label = _WEEK_LABELS.get(str(_wk), f'Week {_wk}')
                    st.markdown(f"### 📅 {_wk_label}")
                    _wk_df = _df_asgn[_df_asgn['Week'] == _wk]
                    for _ck in [k for k in _Y9R_CAMP_LABELS if k in _wk_df['Camp Key'].values]:
                        _camp_df = _wk_df[_wk_df['Camp Key'] == _ck]
                        _has_b = 'B' in _camp_df['Group'].values
                        _camp_label = _Y9R_CAMP_LABELS.get(_ck, _ck)
                        if _has_b:
                            _ca = _camp_df[_camp_df['Group'] == 'A']['Student'].tolist()
                            _cb = _camp_df[_camp_df['Group'] == 'B']['Student'].tolist()
                            _col_a, _col_b = st.columns(2)
                            with _col_a:
                                st.markdown(f"**{_camp_label} — Group A** ({len(_ca)} students)")
                                st.dataframe(pd.DataFrame({'Student': _ca}), hide_index=True, use_container_width=True)
                            with _col_b:
                                st.markdown(f"**{_camp_label} — Group B** ({len(_cb)} students)")
                                st.dataframe(pd.DataFrame({'Student': _cb}), hide_index=True, use_container_width=True)
                        else:
                            _sa = _camp_df['Student'].tolist()
                            st.markdown(f"**{_camp_label}** ({len(_sa)} students)")
                            st.dataframe(pd.DataFrame({'Student': _sa}), hide_index=True, use_container_width=True)

                # ── Excel public roster export ─────────────────────────────────────────────
                _y9r_out = io.BytesIO()
                _wb_y9r = Workbook()

                _header_font  = Font(size=13, bold=True, color="FFFFFF")
                _camp_font    = Font(size=11, bold=True, color="FFFFFF")
                _week_fill    = PatternFill(start_color="FF333333", end_color="FF333333", fill_type="solid")
                _hdr_fill     = PatternFill(start_color="FF444444", end_color="FF444444", fill_type="solid")
                _name_font    = Font(size=10)
                _center       = Alignment(horizontal="center", vertical="center")
                _left         = Alignment(horizontal="left",   vertical="center")
                _thin         = Border(left=Side(style='thin'), right=Side(style='thin'),
                                       top=Side(style='thin'), bottom=Side(style='thin'))

                for _wk in sorted(_df_asgn['Week'].unique()):
                    _ws_name = f"Week {_wk}"
                    _ws = _wb_y9r.create_sheet(title=_ws_name)
                    _wk_label = _WEEK_LABELS.get(str(_wk), f'Week {_wk}')

                    _wk_df = _df_asgn[_df_asgn['Week'] == _wk]
                    _active_camps = [k for k in _Y9R_CAMP_LABELS if k in _wk_df['Camp Key'].values]

                    # Build column layout: each camp (or camp A+B pair) gets columns
                    _col_headers = []  # list of (camp_key, group, col_index_1based)
                    _col_idx = 1
                    for _ck in _active_camps:
                        _cdf = _wk_df[_wk_df['Camp Key'] == _ck]
                        _has_b = 'B' in _cdf['Group'].values
                        if _has_b:
                            _col_headers.append((_ck, 'A', _col_idx)); _col_idx += 1
                            _col_headers.append((_ck, 'B', _col_idx)); _col_idx += 1
                        else:
                            _col_headers.append((_ck, 'A', _col_idx)); _col_idx += 1
                    _total_cols = _col_idx - 1

                    # Row 1: Week header spanning all columns
                    _ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=_total_cols)
                    _wh = _ws.cell(row=1, column=1, value=_wk_label)
                    _wh.font = _header_font; _wh.fill = _week_fill; _wh.alignment = _center

                    # Row 2: Camp headers
                    _prev_ck = None
                    _group_start = None
                    for _ck, _sg, _ci in _col_headers:
                        _c = _ws.cell(row=2, column=_ci)
                        _label = _Y9R_CAMP_LABELS.get(_ck, _ck)
                        _sg_suffix = f" — Group {_sg}" if any(h[0] == _ck and h[1] == 'B' for h in _col_headers) else ""
                        _c.value = _label + _sg_suffix
                        _c.font = Font(bold=True, color="FFFFFF", size=10)
                        _fill_color = _Y9R_COLORS.get(_ck, 'FF555555')
                        _c.fill = PatternFill(start_color=_fill_color, end_color=_fill_color, fill_type="solid")
                        _c.alignment = _center; _c.border = _thin
                        _ws.column_dimensions[get_column_letter(_ci)].width = 26

                    # Rows 3+: Student names
                    _col_lists = {}
                    for _ck, _sg, _ci in _col_headers:
                        _cdf = _wk_df[(_wk_df['Camp Key'] == _ck) & (_wk_df['Group'] == _sg)]
                        _col_lists[_ci] = _cdf['Student'].tolist()

                    _max_rows = max((len(v) for v in _col_lists.values()), default=0)
                    for _r in range(_max_rows):
                        for _ck, _sg, _ci in _col_headers:
                            _lst = _col_lists[_ci]
                            _val = _lst[_r] if _r < len(_lst) else ""
                            _cell = _ws.cell(row=3 + _r, column=_ci, value=_val)
                            _cell.font = _name_font; _cell.alignment = _left; _cell.border = _thin

                # Remove default empty sheet
                if 'Sheet' in _wb_y9r.sheetnames:
                    del _wb_y9r['Sheet']

                _wb_y9r.save(_y9r_out)
                _y9r_out.seek(0)

                st.markdown("---")
                st.download_button(
                    label="📥 Download Y9 Public Roster (Excel)",
                    data=_y9r_out,
                    file_name="Y9_Public_Roster.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# =========================================================================================
# ========================= PAGE 3: Y9 JOURNEY GROUPS ====================================
# =========================================================================================
elif page == "🏔️ Y9 Journey Groups":
    st.title("🏔️ Year 9 Journey Group Sorter")

    # ── Camp Definitions ──────────────────────────────────────────────────────────────────
    Y9_CAMP_DEFS = {
        'MTB':        {'label': 'Mountain to Sea MTB',                                   'max_per': 13},
        'CC':         {'label': 'Colossal Cliffs',                                       'max_per': 12},
        'MM':         {'label': 'Mersey & Mountains',                                    'max_per': 12},
        'Explorer':   {'label': 'Cradle to Coast Explorer (more flexibility)',            'max_per': 16},
        'Challenger': {'label': 'Cradle to Coast Challenger (more physical challenge)',   'max_per': 10},
    }

    Y9_NAME_MAP = {
        'mountain to sea mtb':                           'MTB',
        'colossal cliffs':                               'CC',
        'mersey and mountains':                          'MM',
        'mersey & mountains':                            'MM',
        'cradle to coast explorer (more flexibility)':   'Explorer',
        'cradle to coast explorer':                      'Explorer',
        'cradle to coast challenger (more physical challenge)': 'Challenger',
        'cradle to coast challenger':                    'Challenger',
    }

    WEEK1_HOUSES = {'unwin', 'hodgkin'}
    WEEK2_HOUSES = {'mather', 'ransome'}
    HOUSE_WEEK = {**{h: 1 for h in WEEK1_HOUSES}, **{h: 2 for h in WEEK2_HOUSES}}

    # ── Default AI weights ────────────────────────────────────────────────────────────────
    # Preference rank → score mapping (exponential: pref 5 is near-impossible)
    # Rank: 1=best … 5=worst
    Y9_PREF_SCORES = {1: 500, 2: 200, 3: -800, 4: -5000, 5: -50000}

    Y9_DEFAULT_WEIGHTS = {
        'w_pref':     1,        # Multiplier on Y9_PREF_SCORES (keep at 1; tune via pref_scores)
        'w_friend':   8000,     # Friend pair in same camp reward (raised: guarantee requesters together)
        'w_cap':      600000,   # Over-capacity penalty (per student over limit)
        'w_force':    1000000,  # Forced camp violated penalty
        'w_must':     1000000,  # Must-go-with split penalty
        'w_sep':      500000,   # Can't-go-with together penalty
        'sg_balance': 15,       # Sub-group size imbalance multiplier
        'sg_cap':     600000,   # Sub-group over-capacity penalty
        'sg_friend':  400,      # Sub-group friend pair reward
        # Gender balance weights (applied per camp/subgroup)
        'w_gender_sole':    150000,  # Exactly 1 of a gender in group (worst)
        'w_gender_duo':      30000,  # Exactly 2 of a gender in group (medium)
        'w_gender_allsame':   8000,  # All same gender (better than sole, worse than mixed)
    }

    # ── Session State ─────────────────────────────────────────────────────────────────────
    for _k, _v in [('y9_sep', []), ('y9_must', []), ('y9_force', {}),
                   ('y9_force_week', {}),
                   ('y9_na', []), ('y9_seed', 0), ('y9_results', {}), ('y9_states', {}),
                   ('y9_include_drafts', False), ('y9_generated', False)]:
        if _k not in st.session_state:
            st.session_state[_k] = _v
    if 'y9_weights' not in st.session_state:
        st.session_state.y9_weights = Y9_DEFAULT_WEIGHTS.copy()

    if not (responses_file and students_file):
        st.info("👈 Upload a **Preference Survey CSV** and **Student List CSV** in the sidebar (Step 1) to begin.")
    else:
        # ── Load & Clean ──────────────────────────────────────────────────────────────────
        responses_file.seek(0); students_file.seek(0)
        df_s = pd.read_csv(students_file)
        df_p = pd.read_csv(responses_file)

        # Student list
        df_s['Email'] = df_s['Email'].astype(str).str.strip().str.lower()
        df_s['Official Name'] = (df_s['Preferred name'].astype(str).str.strip()
                                 + " " + df_s['Surname'].astype(str).str.strip())
        df_s['Gender'] = df_s['Gender'].astype(str).str.strip().str.lower() if 'Gender' in df_s.columns else 'o'
        if 'Code' in df_s.columns:
            df_s['Student ID'] = df_s['Code'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
        else:
            df_s['Student ID'] = "N/A"

        # House from student list (fallback for non-responders)
        _house_col_stud = next((c for c in df_s.columns if c.strip().lower() == 'house'), None)
        _rg_col = next((c for c in df_s.columns if 'rollgroup' in c.lower()), None)
        _letter_map = {'u': 'unwin', 'h': 'hodgkin', 'm': 'mather', 'r': 'ransome'}
        _full_house_names = ['unwin', 'hodgkin', 'mather', 'ransome']
        def _house_from_rg(val):
            v = str(val).strip().lower()
            for h in _full_house_names:  # full name anywhere e.g. "9Unwin"
                if h in v: return h
            for ch in ([v[0], v[-1]] if len(v) > 1 else [v[0]] if v else []):
                if ch in _letter_map: return _letter_map[ch]
            return ''
        if _house_col_stud:
            df_s['House_stud'] = df_s[_house_col_stud].astype(str).str.strip().str.lower()
        elif _rg_col:
            df_s['House_stud'] = df_s[_rg_col].apply(_house_from_rg)
        else:
            df_s['House_stud'] = ''

        # Prefs CSV
        df_p['Email address'] = df_p['Email address'].astype(str).str.strip().str.lower()

        # Rank columns: look for "[1]"…"[5]" in header
        _rank_cols = {}
        for _col in df_p.columns:
            for _i in range(1, 6):
                if f'[{_i}]' in _col and ('prefer' in _col.lower() or 'journey' in _col.lower()):
                    _rank_cols[_i] = _col; break

        # Friend columns
        _friend_cols = [c for c in df_p.columns if 'suggest one person' in c.lower()]

        # House column in prefs
        _house_col_pref = next(
            (c for c in df_p.columns if c.strip().lower() == 'which house are you in?'), None)
        if not _house_col_pref:
            _house_col_pref = next((c for c in df_p.columns if 'house' in c.lower()), None)

        # Skill columns
        _camping_col   = next((c for c in df_p.columns if 'general camping skills' in c.lower()), None)
        _sleeping_col  = next((c for c in df_p.columns if 'sleeping outdoors' in c.lower()), None)
        _overnight_col = next((c for c in df_p.columns if 'overnight hike' in c.lower()), None)
        _hardship_col  = next((c for c in df_p.columns if 'tiring' in c.lower() and 'weather' in c.lower()), None)
        _swim_col      = next((c for c in df_p.columns if 'swimming' in c.lower() and 'deep water' in c.lower()), None)
        _challenge_col = next((c for c in df_p.columns if 'physical challenge' in c.lower()), None)
        _ww_col        = next((c for c in df_p.columns if 'white water' in c.lower()), None)
        _teamwork_col  = next((c for c in df_p.columns if 'helping your group' in c.lower() or 'group teamwork' in c.lower()), None)

        # Merge student list with prefs
        df_y9 = pd.merge(
            df_s[['Email', 'Official Name', 'Gender', 'Student ID', 'House_stud']],
            df_p, left_on='Email', right_on='Email address', how='left'
        )
        df_y9['Responded'] = df_y9['Email address'].notna()

        # Determine house and week
        def _get_house(row):
            if row['Responded'] and _house_col_pref:
                h = str(row.get(_house_col_pref, '')).strip().lower()
                if h in HOUSE_WEEK: return h
            h = str(row.get('House_stud', '')).strip().lower()
            return h if h in HOUSE_WEEK else None

        df_y9['House'] = df_y9.apply(_get_house, axis=1)
        df_y9['Week'] = df_y9['House'].map(HOUSE_WEEK)

        # Parse camp preference rankings
        official_names_y9 = df_s['Official Name'].tolist()

        def _parse_prefs(row):
            prefs = {}
            if not row['Responded']: return prefs
            for rank_num, col in _rank_cols.items():
                val = str(row.get(col, '')).strip().lower()
                for frag, key in Y9_NAME_MAP.items():
                    if frag in val:
                        if key not in prefs: prefs[key] = rank_num
                        break
            return prefs

        df_y9['Camp Prefs'] = df_y9.apply(_parse_prefs, axis=1)

        # Friend name matching
        def _match_friend(raw):
            if not raw or not isinstance(raw, str): return None
            raw = raw.strip()
            if raw.lower() in ('', 'nan', 'none'): return None
            raw_l = raw.lower()
            for name in official_names_y9:
                if name.lower() == raw_l: return name
            raw_words = set(raw_l.split())
            for name in official_names_y9:
                if raw_words and raw_words.issubset(set(name.lower().split())): return name
            return None

        def _get_friends(row):
            friends = []
            if not row['Responded']: return friends
            for fc in _friend_cols:
                m = _match_friend(row.get(fc))
                if m and m != row['Official Name'] and m not in friends:
                    friends.append(m)
            return friends

        df_y9['Friends Requested'] = df_y9.apply(_get_friends, axis=1)

        # ── Summary metrics ───────────────────────────────────────────────────────────────
        all_y9 = sorted(df_y9['Official Name'].dropna().unique().tolist())

        # ── SIDEBAR (only visible on this page) ───────────────────────────────────────────
        st.sidebar.markdown("---")

        # ── Load Rules & Weights ──────────────────────────────────────────────────────────
        st.sidebar.header("📂 Load Rules & Weights")
        _y9_json_upload = st.sidebar.file_uploader("Load Saved Y9 JSON", type=["json"], key="y9_json_upload")
        if _y9_json_upload is not None:
            if st.session_state.get('y9_last_loaded') != _y9_json_upload.file_id:
                try:
                    _loaded = json.load(_y9_json_upload)
                    st.session_state.y9_sep        = [tuple(x) for x in _loaded.get('sep',   [])]
                    st.session_state.y9_must       = [tuple(x) for x in _loaded.get('must',  [])]
                    st.session_state.y9_force      = _loaded.get('force', {})
                    st.session_state.y9_force_week = _loaded.get('force_week', {})
                    if 'weights' in _loaded:
                        st.session_state.y9_weights.update(_loaded['weights'])
                    st.session_state.y9_include_drafts = _loaded.get('include_drafts', False)
                    # Resolve not-attending names against the current student list.
                    # Match by exact name, then case-insensitive, then student ID.
                    _all_y9_set    = set(all_y9)
                    _lower_map     = {n.lower(): n for n in all_y9}
                    _id_map        = {
                        str(r.get('Student ID', '')).strip(): r['Official Name']
                        for _, r in df_y9.iterrows()
                    }
                    _l_na = []
                    for _nm in _loaded.get('na', []):
                        _s = str(_nm).strip()
                        if _s in _all_y9_set:
                            _l_na.append(_s)
                        elif _s.lower() in _lower_map:
                            _l_na.append(_lower_map[_s.lower()])
                    for _det in _loaded.get('na_details', []):
                        _sid = str(_det.get('student_id', '')).strip()
                        if _sid in _id_map and _id_map[_sid] not in _l_na:
                            _l_na.append(_id_map[_sid])
                    # Write directly into the widget key — this is what Streamlit
                    # actually reads when it renders the multiselect, identical to
                    # the user having selected those values manually.
                    st.session_state['y9_na_ms'] = _l_na
                    st.session_state.y9_na = _l_na
                    st.session_state.y9_last_loaded = _y9_json_upload.file_id
                    st.sidebar.success("✅ Y9 settings loaded!")
                    st.rerun()
                except Exception as _e:
                    st.sidebar.error(f"❌ Could not parse the JSON file: {_e}")

        st.sidebar.markdown("---")

        # ── Not Attending ─────────────────────────────────────────────────────────────────
        # The widget key 'y9_na_ms' is the single source of truth.
        # On load, we write directly into it above so it behaves exactly
        # like a manual selection — no default= fighting session state.
        st.sidebar.header("🏔️ Y9: Not Attending")
        _y9_na = st.sidebar.multiselect(
            "Students not attending:", options=all_y9, key="y9_na_ms")
        st.session_state.y9_na = _y9_na

        df_y9_act = df_y9[~df_y9['Official Name'].isin(_y9_na)].copy()
        attending_y9 = sorted(df_y9_act['Official Name'].dropna().unique().tolist())

        # Force camp
        st.sidebar.header("📍 Y9: Force Camp")
        _y9_fs = st.sidebar.selectbox("Student", [""] + attending_y9, key="y9_frc_s")
        _y9_fc = st.sidebar.selectbox(
            "Camp", [''] + list(Y9_CAMP_DEFS.keys()),
            format_func=lambda k: Y9_CAMP_DEFS[k]['label'] if k else "-- select --",
            key="y9_frc_c")
        if st.sidebar.button("Add Force Rule", key="y9_add_force"):
            if _y9_fs and _y9_fc:
                st.session_state.y9_force[_y9_fs] = _y9_fc; st.rerun()
        if st.session_state.y9_force:
            st.sidebar.write("**Active Force Rules:**")
            for _s, _c in list(st.session_state.y9_force.items()):
                _c1, _c2 = st.sidebar.columns([5, 1])
                _c1.write(f"- {_s} ➡️ {Y9_CAMP_DEFS.get(_c, {}).get('label', _c)[:22]}")
                if _c2.button("❌", key=f"y9_df_{_s}"):
                    del st.session_state.y9_force[_s]; st.rerun()
            if st.sidebar.button("Clear Force Rules", key="y9_clr_force"):
                st.session_state.y9_force = {}; st.rerun()

        # Force week (opposite-week override)
        st.sidebar.header("🔀 Y9: Force to Opposite Week")
        st.sidebar.caption("Use this when a student must go in the opposite week to their house. E.g. a Unwin/Hodgkin student going with Mather/Ransome.")
        _y9_fw_s = st.sidebar.selectbox("Student", [""] + attending_y9, key="y9_fwk_s")
        _y9_fw_w = st.sidebar.selectbox("Send to Week", ["", 1, 2],
            format_func=lambda w: "-- select --" if w == "" else (
                "Week 1 (Unwin & Hodgkin)" if w == 1 else "Week 2 (Mather & Ransome)"),
            key="y9_fwk_w")
        if st.sidebar.button("Add Force Week Rule", key="y9_add_fwk"):
            if _y9_fw_s and _y9_fw_w != "":
                st.session_state.y9_force_week[_y9_fw_s] = int(_y9_fw_w); st.rerun()
        if st.session_state.y9_force_week:
            st.sidebar.write("**Active Force Week Rules:**")
            for _s, _w in list(st.session_state.y9_force_week.items()):
                _c1, _c2 = st.sidebar.columns([5, 1])
                _wlbl = "Wk 1" if _w == 1 else "Wk 2"
                _c1.write(f"- {_s} ➡️ {_wlbl}")
                if _c2.button("❌", key=f"y9_dfw_{_s}"):
                    del st.session_state.y9_force_week[_s]; st.rerun()
            if st.sidebar.button("Clear Force Week Rules", key="y9_clr_fwk"):
                st.session_state.y9_force_week = {}; st.rerun()

        # Must go with
        st.sidebar.header("🤝 Y9: Must Go With")
        _cm1, _cm2 = st.sidebar.columns(2)
        _y9_m1 = _cm1.selectbox("A", [""] + attending_y9, key="y9_m1")
        _y9_m2 = _cm2.selectbox("B", [""] + attending_y9, key="y9_m2")
        if st.sidebar.button("Add Must-Go Rule", key="y9_add_must"):
            if _y9_m1 and _y9_m2 and _y9_m1 != _y9_m2:
                st.session_state.y9_must.append((_y9_m1, _y9_m2)); st.rerun()
        if st.session_state.y9_must:
            st.sidebar.write("**Active Must-Go Rules:**")
            for _i, _pair in enumerate(st.session_state.y9_must):
                _c1, _c2 = st.sidebar.columns([5, 1])
                _c1.write(f"- {_pair[0]} & {_pair[1]}")
                if _c2.button("❌", key=f"y9_dm_{_i}"):
                    st.session_state.y9_must.pop(_i); st.rerun()
            if st.sidebar.button("Clear Must-Go Rules", key="y9_clr_must"):
                st.session_state.y9_must = []; st.rerun()

        # Can't go with
        st.sidebar.header("⚠️ Y9: Can't Go With")
        _cs1, _cs2 = st.sidebar.columns(2)
        _y9_s1 = _cs1.selectbox("S1", [""] + attending_y9, key="y9_s1")
        _y9_s2 = _cs2.selectbox("S2", [""] + attending_y9, key="y9_s2")
        if st.sidebar.button("Add Can't-Go Rule", key="y9_add_sep"):
            if _y9_s1 and _y9_s2 and _y9_s1 != _y9_s2:
                st.session_state.y9_sep.append((_y9_s1, _y9_s2)); st.rerun()
        if st.session_state.y9_sep:
            st.sidebar.write("**Active Can't-Go Rules:**")
            for _i, _pair in enumerate(st.session_state.y9_sep):
                _c1, _c2 = st.sidebar.columns([5, 1])
                _c1.write(f"- {_pair[0]} ≠ {_pair[1]}")
                if _c2.button("❌", key=f"y9_ds_{_i}"):
                    st.session_state.y9_sep.pop(_i); st.rerun()
            if st.sidebar.button("Clear Can't-Go Rules", key="y9_clr_sep"):
                st.session_state.y9_sep = []; st.rerun()

        st.sidebar.markdown("---")

        # ── Advanced AI Weights ───────────────────────────────────────────────────────────
        with st.sidebar.expander("⚖️ Advanced AI Weights"):
            st.write("Tune how strongly the AI rewards or penalises each factor.")
            _yw = st.session_state.y9_weights
            _yw['w_pref']     = st.number_input("Preference Score Weight",       value=_yw['w_pref'],     step=10,    key="yw_pref")
            _yw['w_friend']   = st.number_input("Friend Pair Reward",            value=_yw['w_friend'],   step=500,   key="yw_friend")
            _yw['w_cap']      = st.number_input("Over-Capacity Penalty",         value=_yw['w_cap'],      step=10000, key="yw_cap")
            _yw['w_force']    = st.number_input("Forced Camp Penalty",           value=_yw['w_force'],    step=10000, key="yw_force")
            _yw['w_must']     = st.number_input("Must-Go-With Penalty",          value=_yw['w_must'],     step=10000, key="yw_must")
            _yw['w_sep']      = st.number_input("Can't-Go-With Penalty",         value=_yw['w_sep'],      step=10000, key="yw_sep")
            _yw['sg_friend']  = st.number_input("Sub-group Friend Reward",       value=_yw['sg_friend'],  step=50,    key="yw_sg_friend")
            st.markdown("**Gender Balance**")
            _yw['w_gender_sole']    = st.number_input("Sole gender in group (1 of a gender) — Penalty",    value=_yw.get('w_gender_sole',    150000), step=10000, key="yw_g_sole")
            _yw['w_gender_duo']     = st.number_input("Duo gender in group (2 of a gender) — Penalty",     value=_yw.get('w_gender_duo',      30000), step=5000,  key="yw_g_duo")
            _yw['w_gender_allsame'] = st.number_input("All same gender in group — Penalty",                value=_yw.get('w_gender_allsame',   8000), step=1000,  key="yw_g_allsame")
            if st.button("↩️ Reset to Defaults", key="y9_reset_wts"):
                st.session_state.y9_weights = Y9_DEFAULT_WEIGHTS.copy(); st.rerun()

        st.sidebar.markdown("---")

        # ── Save Your Setup ───────────────────────────────────────────────────────────────
        st.sidebar.header("💾 Save Your Setup")
        _na_details = []
        for _na_name in st.session_state.y9_na:
            _na_row = df_y9[df_y9['Official Name'] == _na_name]
            if not _na_row.empty:
                _nr = _na_row.iloc[0]
                _na_details.append({
                    'name': _na_name,
                    'student_id': str(_nr.get('Student ID', '')),
                    'house': str(_nr.get('House_stud', '')),
                    'email': str(_nr.get('Email', '')),
                })
        _y9_save = {
            'sep':        [list(p) for p in st.session_state.y9_sep],
            'must':       [list(p) for p in st.session_state.y9_must],
            'force':      st.session_state.y9_force,
            'force_week': st.session_state.y9_force_week,
            'na':         st.session_state.y9_na,
            'na_details': _na_details,
            'include_drafts': st.session_state.y9_include_drafts,
            'weights':    st.session_state.y9_weights,
        }
        st.sidebar.download_button(
            label="📥 Download Y9 Rules & Weights (JSON)",
            data=json.dumps(_y9_save, indent=2),
            file_name="y9_camp_rules.json",
            mime="application/json",
            use_container_width=True,
            key="y9_save_btn"
        )

        # ── MAIN CONTROLS ─────────────────────────────────────────────────────────────────
        st.markdown("---")
        _col_find, _col_ctrl, _col_draft = st.columns([2, 1, 1])
        with _col_find:
            st.subheader("🔍 Find a Student")
            y9_find = st.selectbox("Search:", [""] + attending_y9, key="y9_find_s")
        with _col_ctrl:
            st.subheader("⚙️ Controls")
            y9_depth = st.select_slider(
                "Search Depth:", ["Fast (1-2s)", "Standard (5-10s)", "Deep Search (20-40s)"],
                value="Standard (5-10s)", key="y9_depth")
            _btn_col1, _btn_col2 = st.columns(2)
            with _btn_col1:
                if st.button("▶️ Generate Groups", use_container_width=True, key="y9_generate",
                             type="primary"):
                    st.session_state.y9_generated = True
                    st.session_state.y9_results = {}
                    st.session_state.y9_states = {}
                    st.rerun()
            with _btn_col2:
                if st.button("🔄 Regenerate", use_container_width=True, key="y9_regen",
                             disabled=not st.session_state.y9_generated):
                    st.session_state.y9_seed += 1; st.rerun()
        with _col_draft:
            st.subheader("✏️ Draft Mode")
            _n_no_resp = len(df_y9_act[~df_y9_act["Responded"]])
            st.caption(f"{_n_no_resp} student(s) haven't responded yet.")
            y9_include_drafts = st.checkbox(
                "Draft in non-responders",
                value=st.session_state.y9_include_drafts, key="y9_drafts_cb",
                help="Places students who haven't filled in the survey into groups. "
                     "They are prioritised alongside anyone who requested them, "
                     "then slotted in to balance group sizes. Shown with a grey background.")
            st.session_state.y9_include_drafts = y9_include_drafts
            if y9_include_drafts:
                st.caption("🔘 Non-responders will be placed and shown in grey.")

        y9_find_result = st.empty()
        st.markdown("---")

        # Optimisation pool: built here after the draft checkbox is rendered
        if y9_include_drafts:
            df_y9_opt = df_y9_act.copy()
        else:
            df_y9_opt = df_y9_act[df_y9_act["Responded"]].copy()

        # ── WEEK SUMMARY ──────────────────────────────────────────────────────────────────
        _force_week_rules = st.session_state.get('y9_force_week', {})
        # Compute effective week counts after force-week overrides
        def _effective_week(row):
            nm = row['Official Name']
            if nm in _force_week_rules:
                return _force_week_rules[nm]
            return row['Week']
        df_y9_act['_eff_week'] = df_y9_act.apply(_effective_week, axis=1)
        _w1_n = len(df_y9_act[df_y9_act['_eff_week'] == 1])
        _w2_n = len(df_y9_act[df_y9_act['_eff_week'] == 2])
        _wX_n = len(df_y9_act[df_y9_act['_eff_week'].isna()])
        _fw_n = len([s for s in _force_week_rules if s in attending_y9])
        _mc1, _mc2, _mc3, _mc4, _mc5 = st.columns(5)
        _mc1.metric("Total Y9", len(df_y9_act))
        _mc2.metric("Week 1 (Unwin + Hodgkin)", _w1_n)
        _mc3.metric("Week 2 (Mather + Ransome)", _w2_n)
        _mc4.metric("⚠️ No Week Assigned", _wX_n)
        _mc5.metric("🔀 Force-Week Overrides", _fw_n)
        if _wX_n:
            with st.expander("⚠️ Students with no house/week assigned — check their data"):
                _unassigned = df_y9_act[df_y9_act['Week'].isna()][['Student ID', 'Official Name', 'Responded']]
                st.dataframe(_unassigned.rename(columns={'Official Name': 'Student'}), hide_index=True)

        # Show a clear banner about non-responders who are being excluded
        _n_excl = len(df_y9_act[~df_y9_act['Responded'] & df_y9_act['Week'].notna()])
        if _n_excl > 0 and not y9_include_drafts:
            st.info(
                f"ℹ️ **{_n_excl} student(s) from the student list haven't responded and are not currently "
                f"being placed into groups.** Tick **'Draft in non-responders'** above to include them."
            )

        # ── ALGORITHM HELPERS ─────────────────────────────────────────────────────────────

        def _determine_config(n_students, pref_data_dict, week_num):
            """Return {camp_key: n_instances} obeying per-week camp count limits.

            Rules (same for both weeks unless noted):
              • MTB:  exactly 1
              • CC:   1 or 2
              • MM:   1 or 2
              • Cradle (Explorer + Challenger combined): exactly 2 instances total
                  — could be 2×Explorer, 2×Challenger, or 1 of each
              • Week 1 total: ≤ 7 camps   Week 2 total: ≤ 6 camps
            """
            MAX_CAMPS = {1: 7, 2: 6}.get(week_num, 7)

            # Count demand for each cradle type
            exp_d = sum(1 for p in pref_data_dict.values() if p.get('Explorer',   5) <= 2)
            cha_d = sum(1 for p in pref_data_dict.values() if p.get('Challenger', 5) <= 2)

            # Cradle split: always 2 instances total; split by demand
            if exp_d >= cha_d:
                if exp_d == 0 and cha_d == 0:
                    cradle_config = {'Explorer': 1, 'Challenger': 1}   # no preference data — split evenly
                elif cha_d == 0:
                    cradle_config = {'Explorer': 2}
                else:
                    cradle_config = {'Explorer': 1, 'Challenger': 1}
            else:
                if exp_d == 0:
                    cradle_config = {'Challenger': 2}
                else:
                    cradle_config = {'Explorer': 1, 'Challenger': 1}

            # Start with the fixed minimum: MTB×1, CC×1, MM×1, plus cradle
            config = {'MTB': 1, 'CC': 1, 'MM': 1, **cradle_config}
            total = sum(config.values())   # 4 or 5 depending on cradle split

            # Count demand for CC and MM to decide whether to run a second instance
            cc_d = sum(1 for p in pref_data_dict.values() if p.get('CC', 5) <= 2)
            mm_d = sum(1 for p in pref_data_dict.values() if p.get('MM', 5) <= 2)

            # Add a second MM first (if demand warrants and we have room)
            if total < MAX_CAMPS and mm_d >= cc_d:
                config['MM'] = 2; total += 1
            # Add a second CC (if room)
            if total < MAX_CAMPS:
                config['CC'] = 2; total += 1
            # If MM wasn't doubled yet, try again
            if total < MAX_CAMPS and config.get('MM', 1) == 1:
                config['MM'] = 2; total += 1

            return config

        def _get_caps(config):
            return {k: Y9_CAMP_DEFS[k]['max_per'] * v for k, v in config.items()}

        def _optimize_assignment(week_df, config, caps, friend_reqs, forced,
                                 must_pairs, sep_pairs, seed_str, depth,
                                 gender_lookup=None, pref_dict_override=None,
                                 friend_must_pairs=None):
            """Phase 1: assign every student to a camp TYPE via hill-climbing.

            Key design decisions
            ────────────────────
            • pref_dict_override  — use the caller-built pref dict (includes draft
                                    inheritance) rather than re-reading week_df.
            • friend_must_pairs   — mutual friend requests and draft-student requests
                                    are treated as must-go-with constraints.
            • Hard cap            — any assignment over capacity returns -inf.
            • Smart seeding       — greedy seed places friend-pair members together
                                    in the same camp before random placement.
            • Targeted moves      — 30 % of moves specifically try to unite separated
                                    friend pairs rather than random exploration.
            """
            active   = list(config.keys())
            students = list(week_df['Official Name'])
            # Use the caller-supplied pref dict (has draft inheritance baked in)
            pref_data = pref_dict_override if pref_dict_override else {
                n: (p if isinstance(p, dict) else {})
                for n, p in zip(week_df['Official Name'], week_df['Camp Prefs'])
            }
            _fmp = friend_must_pairs or []
            # Combine explicit must_pairs with friend-derived must pairs
            _all_must = list(must_pairs) + _fmp

            _yw = st.session_state.y9_weights
            W_PREF      = _yw.get('w_pref',   1)
            W_FRIEND    = _yw.get('w_friend',  8000)
            W_FORCE     = _yw.get('w_force',   1000000)
            W_MUST      = _yw.get('w_must',    1000000)
            W_SEP       = _yw.get('w_sep',     500000)
            W_G_SOLE    = _yw.get('w_gender_sole',    150000)
            W_G_DUO     = _yw.get('w_gender_duo',      30000)
            W_G_ALLSAME = _yw.get('w_gender_allsame',   8000)
            _PREF_SCORES = {1: 500, 2: 200, 3: -800, 4: -5000, 5: -50000}

            def _count(asgn):
                c = {x: 0 for x in active}
                for v in asgn.values(): c[v] += 1
                return c

            def _score(asgn):
                s = 0
                for _c, _n in _count(asgn).items():
                    if _n > caps[_c]: return -float('inf')   # hard cap
                for _st, _camp in asgn.items():
                    prefs = pref_data.get(_st, {})
                    s += _PREF_SCORES.get(prefs.get(_camp, 3), -800) * W_PREF
                    for _f in friend_reqs.get(_st, []):
                        if asgn.get(_f) == _camp: s += W_FRIEND
                    if _st in forced and forced[_st] in active:
                        if forced[_st] != _camp: s -= W_FORCE
                for _a, _b in _all_must:
                    if _a in asgn and _b in asgn and asgn[_a] != asgn[_b]: s -= W_MUST
                for _a, _b in sep_pairs:
                    if _a in asgn and _b in asgn and asgn[_a] == asgn[_b]: s -= W_SEP
                if gender_lookup:
                    for _camp in active:
                        _members = [st for st, c in asgn.items() if c == _camp]
                        _m = sum(1 for st in _members if gender_lookup.get(st, 'o') == 'm')
                        _f = sum(1 for st in _members if gender_lookup.get(st, 'o') == 'f')
                        if _m == 0 and _f > 0: s -= W_G_ALLSAME
                        elif _f == 0 and _m > 0: s -= W_G_ALLSAME
                        else:
                            if _m == 1: s -= W_G_SOLE
                            elif _m == 2: s -= W_G_DUO
                            if _f == 1: s -= W_G_SOLE
                            elif _f == 2: s -= W_G_DUO
                return s

            def _greedy_seed(rng):
                """Seed by placing friend/must-pair groups together first."""
                curr = {}
                load = {c: 0 for c in active}

                def _place(st, camp=None):
                    if st in curr: return
                    if camp and load[camp] < caps[camp]:
                        chosen = camp
                    else:
                        prefs  = pref_data.get(st, {})
                        camps_by_pref = sorted(active, key=lambda c: prefs.get(c, 3))
                        chosen = next((c for c in camps_by_pref if load[c] < caps[c]),
                                      min(active, key=lambda c: load[c]))
                    curr[st] = chosen; load[chosen] += 1

                # Forced placements first
                for _st in students:
                    if _st in forced and forced[_st] in active:
                        if load[forced[_st]] < caps[forced[_st]]:
                            _place(_st, forced[_st])

                # Must-pairs (including friend-must-pairs) — place both together
                shuffled_must = list(_all_must); rng.shuffle(shuffled_must)
                for _a, _b in shuffled_must:
                    if _a in students and _b in students:
                        if _a not in curr and _b not in curr:
                            prefs_a = pref_data.get(_a, {})
                            best_c  = sorted(active, key=lambda c: prefs_a.get(c, 3))
                            target  = next((c for c in best_c
                                            if load[c] + 2 <= caps[c]), None)
                            if target:
                                _place(_a, target); _place(_b, target)
                        elif _a in curr and _b not in curr:
                            _place(_b, curr[_a])
                        elif _b in curr and _a not in curr:
                            _place(_a, curr[_b])

                # Remaining students by preference
                remaining = [s for s in students if s not in curr]
                rng.shuffle(remaining)
                for _st in remaining:
                    _place(_st)
                return curr

            if "Fast" in depth:   restarts, iters = 30,  3000
            elif "Standard" in depth: restarts, iters = 120, 6000
            else:                     restarts, iters = 350, 12000

            rng = random.Random(seed_str)
            best_asgn, best_s = None, -float('inf')

            # Build a quick lookup of separated friend pairs for targeted moves
            _separated_pairs = [
                (_a, _b) for _a, _b in _all_must
                if _a in students and _b in students
            ]

            for _ in range(restarts):
                curr   = _greedy_seed(rng)
                curr_s = _score(curr)

                for _i in range(iters):
                    new = dict(curr)
                    r   = rng.random()

                    if r < 0.30 and _separated_pairs:
                        # Targeted move: pick a must/friend pair that is split and
                        # swap one of them to join the other
                        _pair = rng.choice(_separated_pairs)
                        _a, _b = _pair
                        if _a in new and _b in new and new[_a] != new[_b]:
                            # Try moving _a to _b's camp via a swap
                            _camp_b = new[_b]
                            _others_in_b = [s for s in students
                                            if new[s] == _camp_b and s != _b]
                            if _others_in_b:
                                _swap = rng.choice(_others_in_b)
                                new[_a], new[_swap] = new[_swap], new[_a]
                    elif r < 0.75 and len(students) >= 2:
                        # Standard swap (keeps counts balanced → never busts cap)
                        _s1, _s2 = rng.sample(students, 2)
                        new[_s1], new[_s2] = curr[_s2], curr[_s1]
                    else:
                        # Single move to a camp with room
                        _st     = rng.choice(students)
                        _counts = _count(new)
                        _others = [c for c in active
                                   if c != new[_st] and _counts[c] < caps[c]]
                        if _others: new[_st] = rng.choice(_others)

                    ns = _score(new)
                    if ns > curr_s: curr = new; curr_s = ns

                if curr_s > best_s: best_s = curr_s; best_asgn = dict(curr)

            return best_asgn

        def _split_subgroups(camp_students, max_per, friend_reqs, seed_str, gender_lookup=None):
            """Phase 2: split a camp's student list into sub-groups A and B."""
            students = list(camp_students)
            n = len(students)
            if n == 0: return [], []
            if n <= max_per: return students, []

            _yw = st.session_state.y9_weights
            SG_BAL    = _yw.get('sg_balance', 15)
            SG_CAP    = _yw.get('sg_cap',     600000)
            SG_FRIEND = _yw.get('sg_friend',  400)
            SG_G_SOLE    = _yw.get('w_gender_sole',    150000) // 2
            SG_G_DUO     = _yw.get('w_gender_duo',      30000) // 2
            SG_G_ALLSAME = _yw.get('w_gender_allsame',   8000) // 2

            def _sg_score(a, b):
                s = -(abs(len(a) - len(b)) ** 2) * SG_BAL
                if len(a) > max_per: s -= SG_CAP * (len(a) - max_per)
                if len(b) > max_per: s -= SG_CAP * (len(b) - max_per)
                for _st in a:
                    for _f in friend_reqs.get(_st, []):
                        if _f in a: s += SG_FRIEND
                for _st in b:
                    for _f in friend_reqs.get(_st, []):
                        if _f in b: s += SG_FRIEND
                # Gender balance per subgroup
                if gender_lookup:
                    for _grp in [a, b]:
                        if len(_grp) == 0: continue
                        _m = sum(1 for st in _grp if gender_lookup.get(st, 'o') == 'm')
                        _f = sum(1 for st in _grp if gender_lookup.get(st, 'o') == 'f')
                        if _m == 0 and _f > 0: s -= SG_G_ALLSAME
                        elif _f == 0 and _m > 0: s -= SG_G_ALLSAME
                        else:
                            if _m == 1: s -= SG_G_SOLE
                            elif _m == 2: s -= SG_G_DUO
                            if _f == 1: s -= SG_G_SOLE
                            elif _f == 2: s -= SG_G_DUO
                return s

            random.seed(seed_str)
            target = n // 2 + n % 2
            shuffled = students.copy(); random.shuffle(shuffled)
            best_a = set(shuffled[:target]); best_b = set(shuffled[target:])
            best_s = _sg_score(best_a, best_b)

            for _ in range(3000):
                na, nb = set(best_a), set(best_b)
                if na and nb and random.random() < 0.8:
                    _s1 = random.choice(sorted(na)); _s2 = random.choice(sorted(nb))
                    na.discard(_s1); na.add(_s2); nb.discard(_s2); nb.add(_s1)
                elif random.random() < 0.5 and na:
                    _st = random.choice(sorted(na)); na.discard(_st); nb.add(_st)
                elif nb:
                    _st = random.choice(sorted(nb)); nb.discard(_st); na.add(_st)
                ns = _sg_score(na, nb)
                if ns > best_s: best_s = ns; best_a, best_b = na, nb

            return sorted(list(best_a)), sorted(list(best_b))

        # ── PER-WEEK OPTIMIZATION ─────────────────────────────────────────────────────────
        all_week_results = {}  # {week_num: {camp_key: {'A': [...], 'B': [...]}}}
        all_week_configs  = {}
        all_friend_reqs   = {}

        if not st.session_state.y9_generated:
            st.info("👆 Configure your rules above, then click **▶️ Generate Groups** to run the optimiser.")
            st.stop()

        # Non-responders with no house assignment: distribute evenly across weeks
        _unknown_week_drafts = (
            df_y9_opt[df_y9_opt['Week'].isna()].copy()
            if y9_include_drafts else pd.DataFrame()
        )
        if not _unknown_week_drafts.empty:
            # Split alphabetically so the split is deterministic
            _ukd_sorted = _unknown_week_drafts.sort_values('Official Name').reset_index(drop=True)
            _ukd_sorted['_assigned_wk'] = [1 if i % 2 == 0 else 2 for i in range(len(_ukd_sorted))]
        else:
            _ukd_sorted = pd.DataFrame()

        for _wk, _wk_label in [(1, "Week 1 (Unwin & Hodgkin)"), (2, "Week 2 (Mather & Ransome)")]:
            _wdf = df_y9_opt[df_y9_opt['Week'] == _wk].copy()
            # Add any unassigned-week draft students allocated to this week
            if not _ukd_sorted.empty:
                _ukd_this_wk = _ukd_sorted[_ukd_sorted['_assigned_wk'] == _wk].drop(columns=['_assigned_wk'])
                if not _ukd_this_wk.empty:
                    _wdf = pd.concat([_wdf, _ukd_this_wk], ignore_index=True)

            # Apply force-week overrides: students whose home week differs from _wk
            # but have been explicitly forced into this week are added here;
            # students forced out of this week are dropped.
            _force_week_rules = st.session_state.get('y9_force_week', {})
            _forced_into_wk   = [s for s, w in _force_week_rules.items() if w == _wk and s in attending_y9]
            _forced_out_of_wk = [s for s, w in _force_week_rules.items() if w != _wk and s in attending_y9]

            # Add forced-in students (may come from the other week's natural cohort)
            for _fi in _forced_into_wk:
                if _fi not in _wdf['Official Name'].values:
                    _fi_row = df_y9_act[df_y9_act['Official Name'] == _fi]
                    if not _fi_row.empty:
                        _wdf = pd.concat([_wdf, _fi_row], ignore_index=True)

            # Remove forced-out students
            _wdf = _wdf[~_wdf['Official Name'].isin(_forced_out_of_wk)]

            if _wdf.empty:
                all_week_results[_wk] = {}; all_week_configs[_wk] = {}; continue

            _week_students = set(_wdf['Official Name'])

            # ── Build friend requests ────────────────────────────────────────────
            # Step 1: responders' own requests (filtered to this week, minus NA)
            _wfr = {}
            for _, _row in _wdf.iterrows():
                _st = _row['Official Name']
                _fr = [f for f in (_row['Friends Requested'] if isinstance(_row['Friends Requested'], list) else [])
                       if f in _week_students and f not in _y9_na]
                for _a, _b in st.session_state.y9_must:
                    if _st == _a and _b in _week_students and _b not in _fr: _fr.append(_b)
                    if _st == _b and _a in _week_students and _a not in _fr: _fr.append(_a)
                for _a, _b in st.session_state.y9_sep:
                    if _st == _a and _b in _fr: _fr.remove(_b)
                    if _st == _b and _a in _fr: _fr.remove(_a)
                _wfr[_st] = _fr

            # Step 2: build pref_dict now so draft inheritance can use it
            _pref_dict = {
                n: (p if isinstance(p, dict) else {})
                for n, p in zip(_wdf['Official Name'], _wdf['Camp Prefs'])
            }

            # Step 3: draft mode — give non-responders reverse requests + inherited prefs
            if y9_include_drafts:
                _non_resp_in_week = {r['Official Name'] for _, r in _wdf.iterrows() if not r['Responded']}
                for _draft_st in _non_resp_in_week:
                    # Everyone who requested this student
                    _reverse = [o for o, reqs in _wfr.items()
                                if _draft_st in reqs and o != _draft_st]
                    for _a, _b in st.session_state.y9_must:
                        if _draft_st == _a and _b in _week_students and _b not in _reverse:
                            _reverse.append(_b)
                        if _draft_st == _b and _a in _week_students and _a not in _reverse:
                            _reverse.append(_a)
                    _wfr[_draft_st] = _reverse          # ← correct key: _draft_st not _st
                    # Inherit the first requester's camp preferences
                    if _reverse and not _pref_dict.get(_draft_st):
                        _req_prefs = _pref_dict.get(_reverse[0], {})
                        if _req_prefs:
                            _pref_dict[_draft_st] = _req_prefs

            # Step 4: convert friend pairs into pseudo-must pairs so the optimiser
            # treats them as near-hard constraints rather than soft rewards.
            # Only pairs where BOTH students requested each other (mutual) or where
            # one is a non-responder (who can't request back) get elevated.
            _non_resp_names = (
                {r['Official Name'] for _, r in _wdf.iterrows() if not r['Responded']}
                if y9_include_drafts else set()
            )
            _friend_must_pairs = []
            _seen_fp = set()
            for _st, _reqs in _wfr.items():
                for _f in _reqs:
                    _pair = tuple(sorted([_st, _f]))
                    if _pair in _seen_fp: continue
                    _seen_fp.add(_pair)
                    _mutual = _f in _wfr and _st in _wfr[_f]
                    _one_is_draft = _st in _non_resp_names or _f in _non_resp_names
                    if _mutual or _one_is_draft:
                        _friend_must_pairs.append((_st, _f))

            all_friend_reqs.update(_wfr)
            _config = _determine_config(len(_wdf), _pref_dict, _wk)
            _caps   = _get_caps(_config)
            all_week_configs[_wk] = _config

            _forced_wk = {k: v for k, v in st.session_state.y9_force.items() if k in _week_students}
            _sep_wk    = [(a, b) for a, b in st.session_state.y9_sep   if a in _week_students or b in _week_students]
            _must_wk   = [(a, b) for a, b in st.session_state.y9_must  if a in _week_students or b in _week_students]

            _state_key = str({
                'config': _config, 'na': sorted(_y9_na),
                'force': sorted(_forced_wk.items()),
                'force_week': sorted(_force_week_rules.items()),
                'sep':   sorted([tuple(sorted(p)) for p in _sep_wk]),
                'must':  sorted([tuple(sorted(p)) for p in _must_wk]),
                'seed': st.session_state.y9_seed, 'depth': y9_depth, 'n': len(_wdf),
                'drafts': y9_include_drafts,
                'g_sole': st.session_state.y9_weights.get('w_gender_sole', 150000),
                'g_duo':  st.session_state.y9_weights.get('w_gender_duo',   30000),
                'g_all':  st.session_state.y9_weights.get('w_gender_allsame', 8000),
            })

            if st.session_state.y9_states.get(_wk) != _state_key:
                st.session_state.y9_results.pop(_wk, None)
                st.session_state.y9_states[_wk] = _state_key

            if _wk not in st.session_state.y9_results:
                with st.spinner(f"⚙️ Optimising {_wk_label} — {len(_wdf)} students across {sum(_config.values())} camps…"):
                    _gender_lk = dict(zip(_wdf['Official Name'], _wdf['Gender']))
                    _asgn = _optimize_assignment(
                        _wdf, _config, _caps, _wfr,
                        _forced_wk, _must_wk, _sep_wk,
                        f"y9_w{_wk}_{st.session_state.y9_seed}", y9_depth,
                        gender_lookup=_gender_lk,
                        pref_dict_override=_pref_dict,
                        friend_must_pairs=_friend_must_pairs,
                    )

                _week_groups = {}
                for _ck, _ni in _config.items():
                    _camp_list = sorted([s for s, c in _asgn.items() if c == _ck])
                    _mp = Y9_CAMP_DEFS[_ck]['max_per']
                    if _ni == 1:
                        _week_groups[_ck] = {'A': _camp_list, 'B': []}
                    else:
                        _ga, _gb = _split_subgroups(
                            _camp_list, _mp, _wfr,
                            f"split_{_ck}_w{_wk}_{st.session_state.y9_seed}",
                            gender_lookup=_gender_lk)
                        _week_groups[_ck] = {'A': _ga, 'B': _gb}

                st.session_state.y9_results[_wk] = _week_groups

            all_week_results[_wk] = st.session_state.y9_results[_wk]

        # ── BUILD FULL LOOKUP (student → week, camp, subgroup) ────────────────────────────
        full_lookup = {}
        for _wk, _wg in all_week_results.items():
            for _ck, _sgs in _wg.items():
                for _sg, _sgl in _sgs.items():
                    for _nm in _sgl:
                        full_lookup[_nm] = (_wk, _ck, _sg)

        # ── FIND STUDENT RESULT ───────────────────────────────────────────────────────────
        if y9_find and y9_find in full_lookup:
            _fw, _fc, _fsg = full_lookup[y9_find]
            _has_b = len(all_week_results.get(_fw, {}).get(_fc, {}).get('B', [])) > 0
            _sg_str = f" — Group {_fsg}" if _has_b else ""
            _wk_str = "Week 1: Unwin & Hodgkin" if _fw == 1 else "Week 2: Mather & Ransome"
            y9_find_result.success(
                f"🎯 **{y9_find}** → {_wk_str} | **{Y9_CAMP_DEFS[_fc]['label']}{_sg_str}**")
        elif y9_find:
            y9_find_result.warning(
                f"⚠️ {y9_find} was not placed — they may have no house/week assigned.")

        # ── COMPUTE ISOLATED STUDENTS ─────────────────────────────────────────────────────
        isolated_y9 = []
        _seen_iso = set()
        # Build a camp-preference lookup so we can explain why friends weren't placed together
        _y9_pref_lk = {
            row['Official Name']: (row['Camp Prefs'] if isinstance(row['Camp Prefs'], dict) else {})
            for _, row in df_y9.iterrows()
        }
        for _nm, (_fw, _fc, _fsg) in full_lookup.items():
            _friends = all_friend_reqs.get(_nm, [])
            if not _friends: continue
            for _f in _friends:
                if full_lookup.get(_f, (None, None, None))[:2] != (_fw, _fc):
                    if _nm not in _seen_iso:
                        _seen_iso.add(_nm)
                        _f_dest = full_lookup.get(_f)
                        # Determine the most likely reason they weren't placed together
                        if not _f_dest:
                            _iso_why = "Friend not placed — check their house/week data"
                        elif _f_dest[0] != _fw:
                            _f_wk_lbl = "Week 1 (Unwin & Hodgkin)" if _f_dest[0] == 1 else "Week 2 (Mather & Ransome)"
                            _iso_why = f"Friend is in a different week ({_f_wk_lbl}) — different house group"
                        elif any(({_nm, _f} == {a, b}) for a, b in st.session_state.y9_sep):
                            _iso_why = "Separation rule applied by staff"
                        elif _f in st.session_state.y9_force and st.session_state.y9_force[_f] != _fc:
                            _forced_lbl = Y9_CAMP_DEFS.get(st.session_state.y9_force[_f], {}).get('label', st.session_state.y9_force[_f])
                            _iso_why = f"Friend staff-forced to {_forced_lbl}"
                        elif _nm in st.session_state.y9_force and st.session_state.y9_force[_nm] != _fc:
                            _iso_why = f"Student staff-forced to {Y9_CAMP_DEFS.get(_fc, {}).get('label', _fc)}"
                        else:
                            _f_prefs = _y9_pref_lk.get(_f, {})
                            _f_top_camp = min(_f_prefs, key=_f_prefs.get) if _f_prefs else None
                            if _f_top_camp and _f_top_camp != _fc:
                                _iso_why = f"Friend's top preference was {Y9_CAMP_DEFS.get(_f_top_camp, {}).get('label', _f_top_camp)}"
                            else:
                                _iso_why = "Group capacity / competing requests — AI couldn't satisfy all preferences"
                        isolated_y9.append({
                            'Week': _fw,
                            'Student': _nm,
                            'Assigned Camp': Y9_CAMP_DEFS[_fc]['label'],
                            'Friend Requested': _f,
                            'Friend\'s Camp': Y9_CAMP_DEFS[_f_dest[1]]['label'] if _f_dest else 'Not placed',
                            'Why': _iso_why,
                        })
                    break

        # ── CAMP CONFIG SUMMARY ───────────────────────────────────────────────────────────
        st.markdown("---")
        st.subheader("📊 Camp Configuration")
        for _wk in [1, 2]:
            _wdf = df_y9_act[df_y9_act['Week'] == _wk]
            if _wdf.empty: continue
            _config = all_week_configs.get(_wk, {})
            _week_groups = all_week_results.get(_wk, {})
            _wl = "Week 1: Unwin & Hodgkin (12–18 Nov)" if _wk == 1 else "Week 2: Mather & Ransome (23–29 Nov)"
            _nc = sum(_config.values())
            with st.expander(f"📅 {_wl} — {len(_wdf)} students, {_nc} camps {'(6-camp plan)' if _nc == 6 else '(7-camp plan)'}", expanded=True):
                _cfg_rows = []
                for _ck, _ni in _config.items():
                    _mp = Y9_CAMP_DEFS[_ck]['max_per']
                    _grps = _week_groups.get(_ck, {})
                    _sz_a = len(_grps.get('A', []))
                    _sz_b = len(_grps.get('B', []))
                    if _ni == 1:
                        _cfg_rows.append({
                            'Camp': Y9_CAMP_DEFS[_ck]['label'], 'Instances': '1',
                            'Max Per Group': str(_mp), 'Assigned': str(_sz_a), 'Group B': '—'})
                    else:
                        _cfg_rows.append({
                            'Camp': Y9_CAMP_DEFS[_ck]['label'], 'Instances': '2',
                            'Max Per Group': str(_mp), 'Group A': str(_sz_a), 'Group B': str(_sz_b)})
                st.dataframe(pd.DataFrame(_cfg_rows), hide_index=True, use_container_width=True)

        # ── RESULTS TABS (one per camp type) ──────────────────────────────────────────────
        st.markdown("---")
        st.subheader("🏕️ Draft Groups")

        _all_active = set()
        for _wg in all_week_results.values(): _all_active.update(_wg.keys())
        _ordered_camps = [k for k in Y9_CAMP_DEFS if k in _all_active]

        if not _ordered_camps:
            st.info("No groups generated yet. Upload both CSV files to begin.")
        else:
            _tabs = st.tabs([Y9_CAMP_DEFS[k]['label'] for k in _ordered_camps])
            _export_sheets = {}   # {sheet_name: df}

            for _tab, _ck in zip(_tabs, _ordered_camps):
                with _tab:
                    _mp = Y9_CAMP_DEFS[_ck]['max_per']

                    for _wk in [1, 2]:
                        _wdf_hdr = df_y9_act[df_y9_act['Week'] == _wk]
                        # Use the full active pool for lookups so non-responders with
                        # unknown/NaN week (assigned via draft mode) are still found
                        _wdf = df_y9_act.copy()
                        if _wdf_hdr.empty and all_week_results.get(_wk, {}) == {}: continue
                        _week_groups = all_week_results.get(_wk, {})

                        _wt = ("📅 **Week 1: Unwin & Hodgkin** (12–18 November)"
                               if _wk == 1 else
                               "📅 **Week 2: Mather & Ransome** (23–29 November)")
                        st.markdown(_wt)

                        if _ck not in _week_groups:
                            st.info(f"This camp is not scheduled for Week {_wk}.")
                            st.markdown("---"); continue

                        _subgroups = _week_groups[_ck]
                        _has_b = len(_subgroups.get('B', [])) > 0
                        _sg_keys = ['A', 'B'] if _has_b else ['A']

                        _pref_w  = {n: p for n, p in zip(_wdf['Official Name'], _wdf['Camp Prefs']) if isinstance(p, dict)}
                        _fr_w    = {n: f for n, f in zip(_wdf['Official Name'], _wdf['Friends Requested']) if isinstance(f, list)}

                        _sg_containers = st.columns(2) if _has_b else [st.container()]

                        for _sg_i, _sg in enumerate(_sg_keys):
                            _sg_list = _subgroups.get(_sg, [])
                            _cont = _sg_containers[_sg_i] if _has_b else _sg_containers[0]

                            with _cont:
                                _gt = (f"**Group {_sg}** — {len(_sg_list)}/{_mp} students"
                                       if _has_b else
                                       f"**{Y9_CAMP_DEFS[_ck]['label']}** — {len(_sg_list)}/{_mp} students")
                                if len(_sg_list) > _mp:
                                    st.markdown(f"🔴 {_gt} *(OVER CAPACITY)*")
                                else:
                                    st.markdown(_gt)

                                _rows = []
                                for _stud in _sg_list:
                                    _sr = _wdf[_wdf['Official Name'] == _stud]
                                    if _sr.empty: continue
                                    _sr = _sr.iloc[0]
                                    _resp = _sr['Responded']
                                    _friends_raw = _fr_w.get(_stud, [])
                                    _pref_rank = _pref_w.get(_stud, {}).get(_ck, '—')
                                    _friend_str = ', '.join(_friends_raw) if _friends_raw else '—'

                                    _rows.append({
                                        'Student ID':         _sr.get('Student ID', 'N/A'),
                                        'Student':            _stud,
                                        'Responded':          'Yes' if _resp else 'No',
                                        'Gender':             str(_sr.get('Gender', 'o')).upper(),
                                        'House':              str(_sr.get('House', '—')).title(),
                                        'Friend Requested':   _friend_str,
                                        'Pref Rank':          _pref_rank,
                                        'Phys Challenge':     _sr.get(_challenge_col, 'N/A') if _challenge_col and _resp else 'N/A',
                                        'Camping Skill':      _sr.get(_camping_col,   'N/A') if _camping_col   and _resp else 'N/A',
                                        'Overnight Hike':     _sr.get(_overnight_col, 'N/A') if _overnight_col and _resp else 'N/A',
                                        'Hardship Response':  _sr.get(_hardship_col,  'N/A') if _hardship_col  and _resp else 'N/A',
                                        'Swimming':           _sr.get(_swim_col,       'N/A') if _swim_col      and _resp else 'N/A',
                                        'White Water':        _sr.get(_ww_col,         'N/A') if _ww_col        and _resp else 'N/A',
                                        'Group Teamwork':     _sr.get(_teamwork_col,   'N/A') if _teamwork_col  and _resp else 'N/A',
                                    })

                                if _rows:
                                    _df_sg = pd.DataFrame(_rows)

                                    def _hl_y9(row, df_ref, wk, ck, find_name, fl, include_drafts=False):
                                        colors = [''] * len(row)
                                        if row['Student'] == find_name:
                                            return ['background-color: #85e085; font-weight: bold; color: black'] * len(row)
                                        if row['Responded'] == 'No':
                                            if include_drafts:
                                                # Draft student — full grey row
                                                return ['background-color: #d9d9d9; color: #555; font-style: italic'] * len(row)
                                            else:
                                                colors[df_ref.columns.get_loc('Responded')] = 'background-color: #ffffcc'

                                        # Pref rank colouring
                                        _pr_idx = df_ref.columns.get_loc('Pref Rank')
                                        try:
                                            _pr = int(row['Pref Rank'])
                                            if _pr == 1:   colors[_pr_idx] = 'background-color: #c8f7c5; font-weight: bold'
                                            elif _pr == 2: colors[_pr_idx] = 'background-color: #e8f8c5'
                                            elif _pr == 4: colors[_pr_idx] = 'background-color: #ffd9a0'
                                            elif _pr == 5: colors[_pr_idx] = 'background-color: #ffb3b3; font-weight: bold'
                                        except: pass

                                        # Friend colouring
                                        _f_idx = df_ref.columns.get_loc('Friend Requested')
                                        _s_idx = df_ref.columns.get_loc('Student')
                                        _f_str = str(row.get('Friend Requested', ''))
                                        if _f_str and _f_str != '—':
                                            _fl_list = [f.strip() for f in _f_str.split(',')]
                                            _all_together = all(
                                                fl.get(f, (None, None, None))[:2] == (wk, ck)
                                                for f in _fl_list)
                                            if not _all_together:
                                                colors[_f_idx] = 'background-color: #ffcccc'
                                                colors[_s_idx] = 'background-color: #ff9900; color: black; font-weight: bold'

                                        # Separation violation
                                        for _a, _b in st.session_state.y9_sep:
                                            if row['Student'] in (_a, _b):
                                                _other = _b if row['Student'] == _a else _a
                                                if fl.get(_other, (None, None, None))[:2] == (wk, ck):
                                                    colors[_s_idx] = 'background-color: #ff4d4d; color: white; font-weight: bold'

                                        # Skill columns: colour numeric values
                                        for _skill_col in ['Camping Skill', 'Overnight Hike', 'Hardship Response', 'Swimming', 'White Water', 'Group Teamwork', 'Phys Challenge']:
                                            if _skill_col in df_ref.columns:
                                                _sc_idx = df_ref.columns.get_loc(_skill_col)
                                                _val = str(row.get(_skill_col, '')).strip()
                                                _m = re.search(r'^\s*([1-9][0-9]?)(?:\D|$)', _val)
                                                if _m:
                                                    _num = int(_m.group(1))
                                                    if 1 <= _num <= 3:   colors[_sc_idx] = 'background-color: #a83232; color: white; font-weight: bold'
                                                    elif 4 <= _num <= 6: colors[_sc_idx] = 'background-color: #ff9933; color: black; font-weight: bold'
                                                    elif _num >= 7:      colors[_sc_idx] = 'background-color: #c8f7c5; color: black'

                                        return colors

                                    _styled = _df_sg.style.apply(
                                        _hl_y9, df_ref=_df_sg, wk=_wk, ck=_ck,
                                        find_name=y9_find, fl=full_lookup,
                                        include_drafts=y9_include_drafts, axis=1)
                                    st.dataframe(_styled, hide_index=True, use_container_width=True)

                                    _sheet_nm = f"{_ck}{'_'+_sg if _has_b else ''}_W{_wk}"[:31]
                                    _export_sheets[_sheet_nm] = _df_sg
                                else:
                                    st.info("No students in this group.")

                        st.markdown("---")

            # ── DASHBOARDS ────────────────────────────────────────────────────────────────
            st.markdown("---")
            _dcol1, _dcol2, _dcol3 = st.columns(3)
            with _dcol1:
                st.write("### ⚠️ At Risk of Isolation")
                st.caption("Students whose requested friend ended up in a different camp.")
                if isolated_y9:
                    st.dataframe(pd.DataFrame(isolated_y9), hide_index=True, use_container_width=True)
                else:
                    st.success("✅ All students with a friend request are together!")

            with _dcol2:
                st.write("### ❌ Missing Responses")
                # Use the full active pool (df_y9_act) so missing responses always appear
                # regardless of draft-mode setting
                _miss_y9 = df_y9_act[~df_y9_act['Responded']][
                    ['Student ID', 'Official Name', 'Email', 'House_stud']].copy()
                _miss_y9.rename(columns={'Official Name': 'Student', 'House_stud': 'House'}, inplace=True)
                _miss_y9['House'] = _miss_y9['House'].str.title()
                if not _miss_y9.empty:
                    st.dataframe(_miss_y9[['Student ID', 'Student', 'House']], hide_index=True, use_container_width=True)
                    st.write(f"**{len(_miss_y9)} student(s) yet to respond**")
                    _email_list = ', '.join(
                        e for e in _miss_y9['Email'].dropna().tolist()
                        if e and e.lower() not in ('', 'nan', 'none'))
                    if _email_list:
                        st.write("**📧 Email list (copy & paste into BCC):**")
                        st.code(_email_list, language=None)
                else:
                    st.success("✅ Everyone has responded!")

            with _dcol3:
                st.write("### 🚫 Not Attending")
                st.caption("Students marked as not attending this camp.")
                if _y9_na:
                    _na_display = df_y9[df_y9['Official Name'].isin(_y9_na)][
                        ['Student ID', 'Official Name', 'House_stud']].copy()
                    _na_display.rename(columns={'Official Name': 'Student', 'House_stud': 'House'}, inplace=True)
                    _na_display['House'] = _na_display['House'].str.title()
                    st.dataframe(_na_display, hide_index=True, use_container_width=True)
                    st.write(f"**{len(_y9_na)} student(s) not attending**")
                else:
                    st.info("No students marked as not attending.")
            if _export_sheets:
                st.markdown("---")

                def _apply_y9_excel_colors(ws, df_data, sheet_name, full_lk, sep_pairs, find_name, include_drafts):
                    """Apply on-screen highlight colours to a Y9 camp group sheet."""
                    # Parse sheet name to get camp key and week:  e.g. MTB_W1, CC_A_W2
                    _parts = sheet_name.split('_')
                    _ck_xl  = _parts[0]
                    _wk_xl  = int(_parts[-1].replace('W', '')) if _parts[-1].startswith('W') else None

                    fill_blue_hdr   = PatternFill("solid", fgColor="4472C4")
                    fill_green      = PatternFill("solid", fgColor="85E085")
                    fill_grey       = PatternFill("solid", fgColor="D9D9D9")
                    fill_yellow     = PatternFill("solid", fgColor="FFFFCC")
                    fill_pink       = PatternFill("solid", fgColor="FFCCCC")
                    fill_orange     = PatternFill("solid", fgColor="FF9900")
                    fill_red        = PatternFill("solid", fgColor="FF4D4D")
                    fill_pref1      = PatternFill("solid", fgColor="C8F7C5")
                    fill_pref2      = PatternFill("solid", fgColor="E8F8C5")
                    fill_pref4      = PatternFill("solid", fgColor="FFD9A0")
                    fill_pref5      = PatternFill("solid", fgColor="FFB3B3")
                    fill_skill_red  = PatternFill("solid", fgColor="A83232")
                    fill_skill_org  = PatternFill("solid", fgColor="FF9933")
                    fill_skill_grn  = PatternFill("solid", fgColor="C8F7C5")
                    font_bold       = Font(bold=True)
                    font_bold_white = Font(bold=True, color="FFFFFF")
                    font_grey_ital  = Font(color="555555", italic=True)

                    cols = list(df_data.columns)
                    # Style header row
                    for ci in range(1, len(cols) + 1):
                        ws.cell(row=1, column=ci).fill = fill_blue_hdr
                        ws.cell(row=1, column=ci).font = Font(bold=True, color="FFFFFF")

                    skill_cols = ['Camping Skill', 'Overnight Hike', 'Hardship Response',
                                  'Swimming', 'White Water', 'Group Teamwork', 'Phys Challenge']

                    for xl_row, (_, row) in enumerate(df_data.iterrows(), start=2):
                        student  = str(row.get('Student', ''))
                        responded = str(row.get('Responded', '')) == 'Yes'

                        # Found student — full green
                        if find_name and student == find_name:
                            for ci in range(1, len(cols) + 1):
                                ws.cell(row=xl_row, column=ci).fill = fill_green
                                ws.cell(row=xl_row, column=ci).font = font_bold
                            continue

                        # Draft (non-responder) — full grey italic
                        if not responded and include_drafts:
                            for ci in range(1, len(cols) + 1):
                                ws.cell(row=xl_row, column=ci).fill = fill_grey
                                ws.cell(row=xl_row, column=ci).font = font_grey_ital
                            continue

                        # Non-responder not in draft mode — yellow Responded cell
                        if not responded and 'Responded' in cols:
                            ws.cell(row=xl_row, column=cols.index('Responded') + 1).fill = fill_yellow

                        # Pref Rank colouring
                        if 'Pref Rank' in cols:
                            try:
                                _pr = int(row['Pref Rank'])
                                _pr_cell = ws.cell(row=xl_row, column=cols.index('Pref Rank') + 1)
                                if   _pr == 1: _pr_cell.fill = fill_pref1; _pr_cell.font = font_bold
                                elif _pr == 2: _pr_cell.fill = fill_pref2
                                elif _pr == 4: _pr_cell.fill = fill_pref4
                                elif _pr == 5: _pr_cell.fill = fill_pref5; _pr_cell.font = font_bold
                            except (ValueError, TypeError):
                                pass

                        # Friend not in same camp — pink friend cell, orange student cell
                        if 'Friend Requested' in cols and 'Student' in cols:
                            _f_str = str(row.get('Friend Requested', ''))
                            if _f_str and _f_str != '—':
                                _fl_list = [f.strip() for f in _f_str.split(',')]
                                _all_together = all(
                                    full_lk.get(f, (None, None, None))[:2] == (_wk_xl, _ck_xl)
                                    for f in _fl_list)
                                if not _all_together:
                                    ws.cell(row=xl_row, column=cols.index('Friend Requested') + 1).fill = fill_pink
                                    s_cell = ws.cell(row=xl_row, column=cols.index('Student') + 1)
                                    s_cell.fill = fill_orange; s_cell.font = font_bold

                        # Separation violation — red student cell
                        if 'Student' in cols:
                            for _a, _b in sep_pairs:
                                if student in (_a, _b):
                                    _other = _b if student == _a else _a
                                    if full_lk.get(_other, (None, None, None))[:2] == (_wk_xl, _ck_xl):
                                        s_cell = ws.cell(row=xl_row, column=cols.index('Student') + 1)
                                        s_cell.fill = fill_red; s_cell.font = font_bold_white

                        # Skill columns — dark red / orange / green by score
                        for sc in skill_cols:
                            if sc in cols:
                                _val = str(row.get(sc, '')).strip()
                                _m = re.search(r'^\s*([1-9][0-9]?)(?:\D|$)', _val)
                                if _m:
                                    _num = int(_m.group(1))
                                    _sc_cell = ws.cell(row=xl_row, column=cols.index(sc) + 1)
                                    if   1 <= _num <= 3: _sc_cell.fill = fill_skill_red; _sc_cell.font = font_bold_white
                                    elif 4 <= _num <= 6: _sc_cell.fill = fill_skill_org; _sc_cell.font = font_bold
                                    elif _num >= 7:      _sc_cell.fill = fill_skill_grn

                _y9_out = io.BytesIO()
                with pd.ExcelWriter(_y9_out, engine='openpyxl') as _writer:
                    # Ordered sheets: by camp type, then week, then subgroup
                    _ordered_sheets = sorted(
                        _export_sheets.keys(),
                        key=lambda nm: (
                            list(Y9_CAMP_DEFS.keys()).index(nm.split('_')[0])
                            if nm.split('_')[0] in Y9_CAMP_DEFS else 99,
                            nm
                        ))
                    for _snm in _ordered_sheets:
                        _df_sh = _export_sheets[_snm]
                        _df_sh.to_excel(_writer, sheet_name=_snm, index=False)
                        _ws = _writer.sheets[_snm]
                        _apply_y9_excel_colors(_ws, _df_sh, _snm, full_lookup,
                                               st.session_state.y9_sep, y9_find,
                                               y9_include_drafts)
                        for _col in _ws.columns:
                            _ws.column_dimensions[_col[0].column_letter].width = min(
                                max((len(str(cell.value)) for cell in _col), default=0) + 2, 45)

                    # Isolation report sheet
                    if isolated_y9:
                        pd.DataFrame(isolated_y9).to_excel(_writer, sheet_name="Isolation Report", index=False)
                        _ws_iso = _writer.sheets["Isolation Report"]
                        for _col in _ws_iso.columns:
                            _ws_iso.column_dimensions[_col[0].column_letter].width = max(
                                (len(str(cell.value)) for cell in _col), default=0) + 2

                    # Not Attending sheet
                    if _y9_na:
                        _na_df = df_y9[df_y9['Official Name'].isin(_y9_na)][
                            ['Student ID', 'Official Name', 'House_stud', 'Responded']].copy()
                        _na_df.rename(columns={'Official Name': 'Student', 'House_stud': 'House'}, inplace=True)
                        _na_df.to_excel(_writer, sheet_name="Not Attending", index=False)
                        _ws_na = _writer.sheets["Not Attending"]
                        for _col in _ws_na.columns:
                            _ws_na.column_dimensions[_col[0].column_letter].width = max(
                                (len(str(cell.value)) for cell in _col), default=0) + 2

                _y9_out.seek(0)
                st.download_button(
                    label="📥 Download Y9 Journey Groups (Excel)",
                    data=_y9_out,
                    file_name="Y9_Journey_Groups.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# =========================================================================================
# ========================= PAGE 4: Y9 FREE TEXT ANALYSER =================================
# =========================================================================================
elif page == "🔍 Y9 Free Text Analyser":

    # ── Ollama local backend ──────────────────────────────────────────────────────────────
    # Uses Ollama (https://ollama.com) which runs entirely on-device via Metal.
    # Ollama is a single binary that installs into ~/bin — no admin rights needed.
    # It exposes a plain HTTP API, so there are zero MLX threading issues.
    import urllib.request as _urllib_req

    # ── Constants ─────────────────────────────────────────────────────────────────────────
    _FT_COL = "Is there anything else we should know when considering your preferences?"
    _FT_TRIVIAL = {
        "no", "nah", "nar", "no thank you", "not really", "n/a", "none",
        "i don't think so", "nothing", "no.", "no!", "nope", "na", "nil",
        "not that i can think of", "nothing else", "nothing really", "nope!"
    }
    _FT_SYSTEM_PROMPT = """You are a data extraction assistant for a school camp preference survey.
Analyse the student response and return ONLY a valid JSON object with the
fields below. Never invent information not present in the text. If a field
does not apply, use "" or [].

{
  "flags": [],
  "extra_friends": [],
  "tent_notes": "",
  "medical_notes": "",
  "permission_notes": "",
  "preference_notes": "",
  "concern_notes": "",
  "equipment_notes": "",
  "logistics_notes": "",
  "other_notes": "",
  "is_irrelevant": false
}

flags is an array from: ["friend_request", "tent_request", "medical", "permission",
"preference", "concern", "equipment", "logistics", "irrelevant"]

extra_friends: names the student mentions wanting to be WITH, beyond their formal
friend-request field entries. Strings only. Do not include names only mentioned
in the context of tent arrangements.

tent_notes: specific tent arrangements: own tent, shared tent with named people,
tents adjacent, group of 3 in one tent. Different from just wanting to be with
someone socially.

medical_notes: injuries, conditions, medications, allergies, physical limitations.

permission_notes: staff have approved an exception — cross-house, specific group
placement, week override.

preference_notes: camp or activity preferences expressed in text beyond the formal
ranking (e.g. "really wants MTB", "wants Dougie's camp", "wants water-based option").

concern_notes: anxiety, nervousness, conditional attendance ("won't go if...",
"not sure I'm going").

equipment_notes: personal equipment: own tent, own bike, etc.

logistics_notes: scheduling or availability issues (not available week 1, going
different week to house, etc.).

other_notes: anything else relevant to camp staff.

is_irrelevant: true if response has NO actionable info for staff e.g. "I like
fishing", "I'm excited", "I like music"

Return ONLY the JSON object. No explanation. No markdown. No other text."""

    # ── Dependency check ──────────────────────────────────────────────────────────────────
    _ft_mlx_ok = False   # reused as "AI backend ready" flag
    _ft_rf_ok  = False
    try:
        from rapidfuzz import process as _rf_process, fuzz as _rf_fuzz
        _ft_rf_ok = True
    except ImportError:
        st.warning("rapidfuzz is not installed. Run: `pip install rapidfuzz`")

    # ── Ollama status check ───────────────────────────────────────────────────────────────
    # Model: gemma4:e4b  (~5.0 GB 4-bit, Google Gemma 4 E4B — 4.5B effective / 8B total params, edge-optimised)
    # Pull with:  ollama pull gemma4:e4b
    _FT_OLLAMA_URL        = "http://localhost:11434"
    _FT_OLLAMA_MODEL      = "gemma4:e4b"
    _FT_OLLAMA_MODEL_SIZE = "~5.0 GB"
    _FT_OLLAMA_MODEL_DESC = "Google Gemma 4 E4B (4.5B effective, 4-bit)"

    def _ft_ollama_running() -> bool:
        """Return True if the Ollama server is reachable."""
        try:
            _urllib_req.urlopen(f"{_FT_OLLAMA_URL}/api/tags", timeout=2)
            return True
        except Exception:
            return False

    def _ft_ollama_model_available(model: str) -> bool:
        """Return True if the given model has been pulled."""
        try:
            import json as _j
            with _urllib_req.urlopen(f"{_FT_OLLAMA_URL}/api/tags", timeout=3) as _r:
                _tags = _j.loads(_r.read())
            return any(m.get("name", "").startswith(model.split(":")[0])
                       for m in _tags.get("models", []))
        except Exception:
            return False

    def _ft_call_ollama(prompt_text: str) -> str:
        """Send one student response to Ollama and return the raw reply text."""
        import json as _j
        _body = _j.dumps({
            "model":  _FT_OLLAMA_MODEL,
            "prompt": f"{_FT_SYSTEM_PROMPT}\n\nStudent response:\n\n{prompt_text}",
            "stream": False,
            "options": {"temperature": 0, "num_predict": 1024},
            "format": "json",          # tells Ollama to constrain output to valid JSON
        }).encode("utf-8")
        _req = _urllib_req.Request(
            f"{_FT_OLLAMA_URL}/api/generate",
            data=_body,
            headers={"Content-Type": "application/json"},
            method="POST"
        )
        with _urllib_req.urlopen(_req, timeout=60) as _resp:
            _data = _j.loads(_resp.read())
        return _data["response"]

    # ── Sidebar: Ollama status ────────────────────────────────────────────────────────────
    st.sidebar.markdown("---")
    st.sidebar.header("🤖 AI Settings")
    st.sidebar.markdown(
        f"**Model:** `{_FT_OLLAMA_MODEL}`  \n"
        f"**Engine:** {_FT_OLLAMA_MODEL_DESC}  \n"
        f"**Size:** {_FT_OLLAMA_MODEL_SIZE}  \n"
        f"**Backend:** Ollama (on-device, Metal)"
    )
    _ollama_up    = _ft_ollama_running()
    _model_ready  = _ollama_up and _ft_ollama_model_available(_FT_OLLAMA_MODEL)

    if _ollama_up and _model_ready:
        st.sidebar.success(f"✅ Ollama running — `{_FT_OLLAMA_MODEL}` ready")
        _ft_mlx_ok = True
    elif _ollama_up and not _model_ready:
        st.sidebar.warning(
            f"⚠️ Ollama is running but **{_FT_OLLAMA_MODEL}** hasn't been downloaded yet.\n\n"
            f"Open Terminal and run:\n```\nollama pull {_FT_OLLAMA_MODEL}\n```"
        )
    else:
        st.sidebar.error(
            "❌ Ollama is not running.\n\n"
            "**One-time setup (no admin needed):**\n"
            "1. Download from [ollama.com](https://ollama.com) and open the `.dmg` — "
            "it installs to your home folder.\n"
            f"2. In Terminal: `ollama pull {_FT_OLLAMA_MODEL}`\n"
            "3. Refresh this page."
        )

    # ── Session state init ─────────────────────────────────────────────────────────────────
    for _ftk, _ftv in [
        ("y9_ft_results",         []),
        ("y9_ft_analysed",        False),
        ("y9_ft_new_must",        []),
        ("y9_ft_new_force_week",  {}),
        ("y9_ft_noted_list",      []),   # list of (email, category) tuples — kept for compat
        ("y9_ft_friend_dec",      {}),   # (email, raw_nm) -> {action, resolved}
        ("y9_ft_perm_dec",        {}),   # (email, 'perm') -> {type, ...}
        ("y9_ft_staff_ticks",     {}),   # (email, field_key) -> bool — which notes to export
        ("y9_ft_manual_add",      {}),   # email -> list of {type, detail} manual additions
    ]:
        if _ftk not in st.session_state:
            st.session_state[_ftk] = _ftv

    # ── Sidebar: existing rules uploader (page-gated) ─────────────────────────────────────
    st.sidebar.markdown("---")
    st.sidebar.header("📂 FT Analyser: Existing Rules")
    _ft_rules_upload = st.sidebar.file_uploader(
        "Load Existing Y9 Rules JSON (optional)",
        type=["json"], key="y9_ft_rules_upload"
    )

    # ── Title ─────────────────────────────────────────────────────────────────────────────
    st.title("🔍 Y9 Free Text Analyser")
    st.caption(
        "Pre-processing tool — runs the AI over free-text survey responses, "
        "surfaces actionable items for review, and exports a combined rules JSON "
        "ready for the Y9 Journey Groups tool."
    )

    # ── Data check ────────────────────────────────────────────────────────────────────────
    if not (responses_file and students_file):
        st.info("👈 Upload both CSV files using **Step 1 — Core Data Upload** in the sidebar to begin.")
        st.stop()

    # ── Load & clean data ─────────────────────────────────────────────────────────────────
    responses_file.seek(0); students_file.seek(0)
    _ft_df_s = pd.read_csv(students_file)
    _ft_df_p = pd.read_csv(responses_file)

    _ft_df_s["Email"] = _ft_df_s["Email"].astype(str).str.strip().str.lower()
    _ft_df_s["Official Name"] = (
        _ft_df_s["Preferred name"].astype(str).str.strip()
        + " " + _ft_df_s["Surname"].astype(str).str.strip()
    )

    # House / week
    _ft_full_houses  = ["unwin", "hodgkin", "mather", "ransome"]
    _ft_letter_map   = {"u": "unwin", "h": "hodgkin", "m": "mather", "r": "ransome"}
    _ft_house_col_s  = next((c for c in _ft_df_s.columns if c.strip().lower() == "house"), None)
    _ft_rg_col       = next((c for c in _ft_df_s.columns if "rollgroup" in c.lower()), None)

    def _ft_house_from_rg(val):
        v = str(val).strip().lower()
        for h in _ft_full_houses:
            if h in v: return h
        for ch in ([v[0], v[-1]] if len(v) > 1 else [v[0]] if v else []):
            if ch in _ft_letter_map: return _ft_letter_map[ch]
        return ""

    if _ft_house_col_s:
        _ft_df_s["House"] = _ft_df_s[_ft_house_col_s].astype(str).str.strip().str.lower()
    elif _ft_rg_col:
        _ft_df_s["House"] = _ft_df_s[_ft_rg_col].apply(_ft_house_from_rg)
    else:
        _ft_df_s["House"] = ""

    _FT_HOUSE_WEEK = {"unwin": 1, "hodgkin": 1, "mather": 2, "ransome": 2}
    _ft_df_s["Week"] = _ft_df_s["House"].map(_FT_HOUSE_WEEK)

    _ft_df_p["Email address"] = _ft_df_p["Email address"].astype(str).str.strip().str.lower()

    _ft_df = pd.merge(
        _ft_df_s[["Email", "Official Name", "House", "Week"]],
        _ft_df_p, left_on="Email", right_on="Email address", how="left"
    )
    _ft_df["Responded"] = _ft_df["Email address"].notna()

    _ft_official_names = sorted(_ft_df_s["Official Name"].dropna().unique().tolist())

    # Formal friend columns
    _ft_friend_cols = [c for c in _ft_df_p.columns if "suggest one person" in c.lower()]

    # Build formal friend lookup: name -> [formally requested official names]
    def _ft_get_formal_friends_for(row):
        out = []
        if pd.isna(row.get("Email address")): return out
        for _fc in _ft_friend_cols:
            val = row.get(_fc)
            if not (pd.notna(val) and isinstance(val, str) and val.strip()): continue
            val_s = val.strip()
            # exact then case-insensitive match
            if val_s in _ft_official_names and val_s not in out:
                out.append(val_s); continue
            for nm in _ft_official_names:
                if val_s.lower() == nm.lower() and nm not in out:
                    out.append(nm); break
        return out

    _ft_formal_friends = {}
    for _, _rw in _ft_df.iterrows():
        _ft_formal_friends[_rw["Official Name"]] = _ft_get_formal_friends_for(_rw)

    def _ft_is_already_mutual(name_a, name_b):
        return (name_b in _ft_formal_friends.get(name_a, [])
                and name_a in _ft_formal_friends.get(name_b, []))

    # Locate free-text column
    _ft_col_actual = next(
        (c for c in _ft_df_p.columns if "anything else we should know" in c.lower()),
        None
    )

    # Pre-filter: build list of rows with non-trivial free text
    def _ft_is_trivial(text):
        if pd.isna(text): return True
        t = str(text).strip()
        if not t: return True
        if t.lower() in _FT_TRIVIAL: return True
        if all(ch in ".,!?;: \t\n-" for ch in t): return True
        return False

    _ft_rows_to_analyse = []
    if _ft_col_actual:
        for _, _rw in _ft_df.iterrows():
            _txt = _rw.get(_ft_col_actual)
            if not _ft_is_trivial(_txt):
                _ft_rows_to_analyse.append({
                    "email": _rw["Email"],
                    "name":  _rw["Official Name"],
                    "house": str(_rw.get("House", "") or "").title(),
                    "week":  _rw.get("Week"),
                    "text":  str(_txt).strip(),
                })
    else:
        st.error(f"Could not find the free-text column. Expected to contain: \"anything else we should know\"")

    # Metrics
    _ft_total_resp = int(_ft_df["Responded"].sum())
    _ft_n_analyse  = len(_ft_rows_to_analyse)
    _ft_n_trivial  = _ft_total_resp - _ft_n_analyse

    _mc1, _mc2, _mc3 = st.columns(3)
    _mc1.metric("Total responses", _ft_total_resp)
    _mc2.metric("Free text to analyse", _ft_n_analyse)
    _mc3.metric("Trivial / blank (skipped)", _ft_n_trivial)

    st.markdown("---")

    # ── Analyse / Reset buttons ───────────────────────────────────────────────────────────
    _ft_btn_col, _ft_rst_col = st.columns([3, 1])
    with _ft_btn_col:
        _ft_analyse_clicked = st.button(
            "🔬 Analyse Free-Text Responses",
            disabled=(not _ft_mlx_ok) or (not _ft_rf_ok) or st.session_state.y9_ft_analysed,
            use_container_width=True,
            type="primary",
            key="y9_ft_analyse_btn",
            help="Start Ollama and pull the model first — see the sidebar." if not _ft_mlx_ok else None
        )
    with _ft_rst_col:
        _ft_reset_clicked = st.button(
            "🗑️ Reset",
            use_container_width=True,
            disabled=not st.session_state.y9_ft_analysed,
            key="y9_ft_reset_btn"
        )

    if _ft_reset_clicked:
        for _rk in ["y9_ft_results", "y9_ft_new_must", "y9_ft_noted_list"]:
            st.session_state[_rk] = []
        for _rk in ["y9_ft_new_force_week", "y9_ft_friend_dec", "y9_ft_perm_dec",
                    "y9_ft_staff_ticks", "y9_ft_manual_add"]:
            st.session_state[_rk] = {}
        st.session_state.y9_ft_analysed = False
        st.rerun()

    # ── Analysis execution ────────────────────────────────────────────────────────────────
    if _ft_analyse_clicked and _ft_mlx_ok and _ft_rf_ok:

        _ft_results   = []
        _ft_n         = len(_ft_rows_to_analyse)
        _ft_prog      = st.progress(0, text="Starting analysis…")

        for _fi, _frow in enumerate(_ft_rows_to_analyse):
            _ft_prog.progress(
                _fi / max(_ft_n, 1),
                text=f"Analysing response {_fi + 1} of {_ft_n} — {_frow['name']}…"
            )

            try:
                _raw = _ft_call_ollama(_frow["text"])

                # Strip accidental markdown fences
                _clean = _raw.strip()
                if _clean.startswith("```"):
                    _clean = re.sub(r"^```(?:json)?\s*", "", _clean)
                    _clean = re.sub(r"\s*```$", "", _clean).strip()
                # Trim after closing brace
                _close = _clean.rfind("}")
                if _close != -1:
                    _clean = _clean[:_close + 1]

                _llm_result = json.loads(_clean)

                # Fuzzy-match extra_friends
                _name_matches = {}
                for _raw_nm in (_llm_result.get("extra_friends") or []):
                    _hits = _rf_process.extract(
                        str(_raw_nm), _ft_official_names,
                        scorer=_rf_fuzz.token_sort_ratio, limit=3
                    )
                    _name_matches[_raw_nm] = [(nm, sc) for nm, sc, _ in _hits if sc >= 65]

                _ft_results.append({
                    "email": _frow["email"], "name": _frow["name"],
                    "house": _frow["house"], "week": _frow["week"],
                    "text":  _frow["text"],  "llm":  _llm_result,
                    "name_matches": _name_matches, "error": False,
                })

            except Exception as _ex:
                _ft_results.append({
                    "email": _frow["email"], "name": _frow["name"],
                    "house": _frow["house"], "week": _frow["week"],
                    "text":  _frow["text"],  "llm":  {},
                    "name_matches": {}, "error": True, "error_msg": str(_ex),
                })

        _ft_prog.progress(1.0, text="✅ Analysis complete!")
        st.session_state.y9_ft_results  = _ft_results
        st.session_state.y9_ft_analysed = True
        st.rerun()

    # ── Review queue ──────────────────────────────────────────────────────────────────────
    if st.session_state.y9_ft_analysed and st.session_state.y9_ft_results:

        _ft_res_all   = st.session_state.y9_ft_results
        _ft_actionable = [r for r in _ft_res_all
                          if not r["error"] and not r["llm"].get("is_irrelevant", False)]
        _ft_errors     = [r for r in _ft_res_all if r["error"]]

        # Sort: medical → permission → friend/tent → concern → other
        def _ft_sort_key(r):
            _fl = set(r["llm"].get("flags", []))
            if "medical"    in _fl: return 0
            if "permission" in _fl: return 1
            if "friend_request" in _fl or "tent_request" in _fl: return 2
            if "concern"    in _fl: return 3
            return 4

        _ft_actionable.sort(key=_ft_sort_key)

        # Summary bar
        _ft_noted_set  = set(tuple(x) for x in st.session_state.y9_ft_noted_list)
        _ft_n_must     = len(st.session_state.y9_ft_new_must)
        _ft_n_med_noted = sum(1 for (_, cat) in _ft_noted_set if cat in ("Medical", "Concern"))
        _ft_n_cards    = len(_ft_actionable)
        _ft_n_irrel    = len([r for r in _ft_res_all if not r["error"] and r["llm"].get("is_irrelevant", False)])

        _sb1, _sb2, _sb3, _sb4 = st.columns(4)
        _sb1.metric("Cards with actionable content", _ft_n_cards)
        _sb2.metric("New Must-Go rules ready", _ft_n_must)
        _sb3.metric("Medical / Concern flags noted", _ft_n_med_noted)
        _sb4.metric("Irrelevant responses (skipped)", _ft_irrel if (_ft_irrel := _ft_n_irrel) else 0)

        st.markdown("---")
        st.subheader("📋 Review Queue")

        _FT_FLAG_EMOJI = {
            "friend_request": "🟠 Friend request",
            "tent_request":   "🔵 Tent",
            "medical":        "🔴 Medical",
            "permission":     "🟡 Permission",
            "preference":     "🟢 Preference",
            "concern":        "🟣 Concern",
            "equipment":      "⚪ Equipment",
            "logistics":      "⚪ Logistics",
            "irrelevant":     "— Irrelevant",
        }

        for _fi, _fres in enumerate(_ft_actionable):
            _fl = _fres["llm"].get("flags", [])
            _fl_labels = "  ".join(_FT_FLAG_EMOJI.get(f, f) for f in _fl if f != "irrelevant")
            _card_label = f"**{_fres['name']}**  —  {_fl_labels}" if _fl_labels else _fres["name"]
            _expand_default = "medical" in _fl or "concern" in _fl

            with st.expander(_card_label, expanded=_expand_default):
                st.info(f"📝 **Original response:** {_fres['text']}")
                _ll = _fres["llm"]

                # ── Extra friend requests ──────────────────────────────────────────────────
                _extra_friends = _ll.get("extra_friends") or []
                if _extra_friends:
                    st.markdown("**🟠 Extra friend mentions (beyond formal fields):**")
                    for _raw_nm in _extra_friends:
                        _matches  = _fres["name_matches"].get(_raw_nm, [])
                        _dec_key  = (_fres["email"], _raw_nm)
                        _decision = st.session_state.y9_ft_friend_dec.get(_dec_key)

                        if _decision:
                            if _decision["action"] == "accept":
                                st.success(f"✅ **{_raw_nm}** → Must-Go rule added: **{_fres['name']}** + **{_decision['resolved']}**")
                            else:
                                st.warning(f"❌ **{_raw_nm}** — Dismissed")
                            continue

                        # Check mutual against top match
                        _top_match = _matches[0][0] if _matches else None
                        if _top_match and _ft_is_already_mutual(_fres["name"], _top_match):
                            st.success(f"✅ **{_raw_nm}** → Already captured in formal fields (mutual: {_fres['name']} ↔ {_top_match})")
                            continue

                        _safe = re.sub(r"\W+", "_", _raw_nm)[:18]
                        _sel_key     = f"y9_ft_sel_{_fi}_{_safe}"
                        _accept_key  = f"y9_ft_acc_{_fi}_{_safe}"
                        _dismiss_key = f"y9_ft_dis_{_fi}_{_safe}"

                        st.markdown(f"*Extracted name:* `{_raw_nm}`")
                        _sc_a, _sc_b = st.columns([3, 2])
                        with _sc_a:
                            if _matches:
                                _opts = [m[0] for m in _matches] + ["— not listed —"]
                                _scores = {m[0]: m[1] for m in _matches}
                                st.selectbox(
                                    f"Best matches for \"{_raw_nm}\":",
                                    _opts,
                                    format_func=lambda n: f"{n} ({_scores[n]:.0f}%)" if n in _scores else n,
                                    key=_sel_key
                                )
                            else:
                                st.caption(f"No confident match for \"{_raw_nm}\" — search manually:")
                                st.selectbox(
                                    "Search roster:",
                                    ["— not listed —"] + _ft_official_names,
                                    key=_sel_key
                                )
                        with _sc_b:
                            _bca, _bcb = st.columns(2)
                            if _bca.button("✅ Must-Go", key=_accept_key, use_container_width=True):
                                _resolved = st.session_state.get(_sel_key, "")
                                if _resolved and _resolved != "— not listed —":
                                    _pair     = (_fres["name"], _resolved)
                                    _rev_pair = (_resolved, _fres["name"])
                                    _existing = [(a, b) for (a, b) in st.session_state.y9_ft_new_must]
                                    if _pair not in _existing and _rev_pair not in _existing:
                                        st.session_state.y9_ft_new_must.append(_pair)
                                    st.session_state.y9_ft_friend_dec[_dec_key] = {
                                        "action": "accept", "resolved": _resolved
                                    }
                                    st.rerun()
                            if _bcb.button("❌ Dismiss", key=_dismiss_key, use_container_width=True):
                                st.session_state.y9_ft_friend_dec[_dec_key] = {
                                    "action": "dismiss", "resolved": None
                                }
                                st.rerun()

                # ── Tent notes ────────────────────────────────────────────────────────────
                _tent_raw = _ll.get("tent_notes")
                if isinstance(_tent_raw, list): _tent_raw = " ".join(str(x) for x in _tent_raw if x)
                if _tent := (str(_tent_raw) if _tent_raw else "").strip():
                    _tent_tick_key = (_fres["email"], "tent_notes")
                    _tent_ticked   = st.session_state.y9_ft_staff_ticks.get(_tent_tick_key, False)
                    _new_tent_tick = st.checkbox(
                        f"🔵 **Tent:** {_tent}",
                        value=_tent_ticked,
                        key=f"y9_ft_tick_tent_{_fi}",
                        help="Tick to include in Staff Notes export"
                    )
                    if _new_tent_tick != _tent_ticked:
                        st.session_state.y9_ft_staff_ticks[_tent_tick_key] = _new_tent_tick
                        st.rerun()

                # ── Medical notes ─────────────────────────────────────────────────────────
                _med_raw = _ll.get("medical_notes")
                if isinstance(_med_raw, list): _med_raw = " ".join(str(x) for x in _med_raw if x)
                if _med := (str(_med_raw) if _med_raw else "").strip():
                    _med_tick_key = (_fres["email"], "medical_notes")
                    _med_ticked   = st.session_state.y9_ft_staff_ticks.get(_med_tick_key, False)
                    _new_med_tick = st.checkbox(
                        f"🔴 **Medical:** {_med}",
                        value=_med_ticked,
                        key=f"y9_ft_tick_med_{_fi}",
                        help="Tick to include in Staff Notes export"
                    )
                    if _new_med_tick != _med_ticked:
                        st.session_state.y9_ft_staff_ticks[_med_tick_key] = _new_med_tick
                        st.rerun()

                # ── Permission notes ──────────────────────────────────────────────────────
                _perm_raw = _ll.get("permission_notes")
                if isinstance(_perm_raw, list): _perm_raw = " ".join(str(x) for x in _perm_raw if x)
                if _perm := (str(_perm_raw) if _perm_raw else "").strip():
                    st.markdown("**🟡 Permission:**")
                    st.write(_perm)
                    _perm_key = (_fres["email"], "perm")
                    _perm_dec = st.session_state.y9_ft_perm_dec.get(_perm_key)
                    if _perm_dec:
                        st.success(f"✅ {_perm_dec['type']}")
                    else:
                        st.caption("Choose the appropriate action:")
                        _pmg_key  = f"y9_ft_perm_partner_{_fi}"
                        _pfw_key  = f"y9_ft_perm_wk_{_fi}"
                        _pc1, _pc2, _pc3 = st.columns([3, 2, 1])
                        with _pc1:
                            st.selectbox(
                                "Must-Go with student:",
                                ["— select —"] + _ft_official_names,
                                key=_pmg_key, label_visibility="collapsed"
                            )
                            if st.button("✅ Add Must-Go Rule", key=f"y9_ft_pmg_{_fi}", use_container_width=True):
                                _partner = st.session_state.get(_pmg_key, "")
                                if _partner and _partner != "— select —":
                                    _pair     = (_fres["name"], _partner)
                                    _rev_pair = (_partner, _fres["name"])
                                    _existing = list(st.session_state.y9_ft_new_must)
                                    if _pair not in _existing and _rev_pair not in _existing:
                                        st.session_state.y9_ft_new_must.append(_pair)
                                    st.session_state.y9_ft_perm_dec[_perm_key] = {
                                        "type": f"Must-Go: {_fres['name']} + {_partner}"
                                    }
                                    st.rerun()
                        with _pc2:
                            st.selectbox(
                                "Force to Week:", [1, 2],
                                format_func=lambda w: f"Week {w}",
                                key=_pfw_key, label_visibility="collapsed"
                            )
                            if st.button("✅ Force-Week Override", key=f"y9_ft_pfw_{_fi}", use_container_width=True):
                                _wk = int(st.session_state.get(_pfw_key, 1))
                                st.session_state.y9_ft_new_force_week[_fres["name"]] = _wk
                                st.session_state.y9_ft_perm_dec[_perm_key] = {
                                    "type": f"Force to Week {_wk}"
                                }
                                st.rerun()
                        with _pc3:
                            if st.button("⏭️ Skip", key=f"y9_ft_pskip_{_fi}", use_container_width=True):
                                st.session_state.y9_ft_perm_dec[_perm_key] = {"type": "No action"}
                                st.rerun()

                # ── Concern notes ─────────────────────────────────────────────────────────
                _concern_raw = _ll.get("concern_notes")
                if isinstance(_concern_raw, list): _concern_raw = " ".join(str(x) for x in _concern_raw if x)
                if _concern := (str(_concern_raw) if _concern_raw else "").strip():
                    _con_tick_key = (_fres["email"], "concern_notes")
                    _con_ticked   = st.session_state.y9_ft_staff_ticks.get(_con_tick_key, False)
                    _new_con_tick = st.checkbox(
                        f"🟣 **Concern:** {_concern}",
                        value=_con_ticked,
                        key=f"y9_ft_tick_concern_{_fi}",
                        help="Tick to include in Staff Notes export"
                    )
                    if _new_con_tick != _con_ticked:
                        st.session_state.y9_ft_staff_ticks[_con_tick_key] = _new_con_tick
                        st.rerun()

                # ── Read-only notes with tick-to-export checkboxes ────────────────────────
                for _note_field, _note_label, _note_emoji in [
                    ("preference_notes", "Preference", "🟢"),
                    ("equipment_notes",  "Equipment",  "⚪"),
                    ("logistics_notes",  "Logistics",  "⚪"),
                    ("other_notes",      "Other",      "📝"),
                ]:
                    _raw_val = _ll.get(_note_field)
                    if isinstance(_raw_val, list): _raw_val = " ".join(str(x) for x in _raw_val if x)
                    if _note_val := (str(_raw_val) if _raw_val else "").strip():
                        _n_tick_key = (_fres["email"], _note_field)
                        _n_ticked   = st.session_state.y9_ft_staff_ticks.get(_n_tick_key, False)
                        _new_n_tick = st.checkbox(
                            f"{_note_emoji} **{_note_label}:** {_note_val}",
                            value=_n_ticked,
                            key=f"y9_ft_tick_{_note_field}_{_fi}",
                            help="Tick to include in Staff Notes export"
                        )
                        if _new_n_tick != _n_ticked:
                            st.session_state.y9_ft_staff_ticks[_n_tick_key] = _new_n_tick
                            st.rerun()

                # ── Manual additions ──────────────────────────────────────────────────────
                st.markdown("---")
                st.markdown("**➕ Manual additions** — add anything the AI missed:")

                _man_email   = _fres["email"]
                _man_entries = st.session_state.y9_ft_manual_add.get(_man_email, [])

                # Show already-added manual entries
                for _mi, _me in enumerate(_man_entries):
                    _mc_label = {
                        "pairing":    "🟠 Pairing",
                        "force_camp": "🏕️ Forced camp",
                        "force_week": "📅 Forced week",
                        "staff_note": "📝 Staff note",
                    }.get(_me["type"], _me["type"])
                    _mcol1, _mcol2 = st.columns([5, 1])
                    _mcol1.success(f"{_mc_label}: **{_me['detail']}**")
                    if _mcol2.button("✕", key=f"y9_ft_man_del_{_man_email}_{_mi}", help="Remove"):
                        _man_entries.pop(_mi)
                        st.session_state.y9_ft_manual_add[_man_email] = _man_entries
                        st.rerun()

                # Add-new form
                _man_type_key    = f"y9_ft_man_type_{_fi}"
                _man_partner_key = f"y9_ft_man_partner_{_fi}"
                _man_camp_key    = f"y9_ft_man_camp_{_fi}"
                _man_week_key    = f"y9_ft_man_week_{_fi}"
                _man_note_key    = f"y9_ft_man_note_{_fi}"

                _man_type = st.selectbox(
                    "Add type:",
                    ["— select —", "pairing", "force_camp", "force_week", "staff_note"],
                    format_func=lambda x: {
                        "— select —":  "— choose type —",
                        "pairing":     "🟠 Student pairing (Must-Go)",
                        "force_camp":  "🏕️ Forced camp assignment",
                        "force_week":  "📅 Forced week override",
                        "staff_note":  "📝 Custom staff note",
                    }.get(x, x),
                    key=_man_type_key,
                    label_visibility="collapsed"
                )

                if _man_type == "pairing":
                    _man_partner = st.selectbox(
                        "Must go with:", ["— select —"] + _ft_official_names,
                        key=_man_partner_key, label_visibility="collapsed"
                    )
                    if st.button("➕ Add Pairing", key=f"y9_ft_man_add_{_fi}_pair", use_container_width=True):
                        if _man_partner and _man_partner != "— select —":
                            _pair     = (_fres["name"], _man_partner)
                            _rev_pair = (_man_partner, _fres["name"])
                            _existing_m = list(st.session_state.y9_ft_new_must)
                            if _pair not in _existing_m and _rev_pair not in _existing_m:
                                st.session_state.y9_ft_new_must.append(_pair)
                            _detail = f"{_fres['name']} + {_man_partner}"
                            _man_entries.append({"type": "pairing", "detail": _detail})
                            st.session_state.y9_ft_manual_add[_man_email] = _man_entries
                            st.rerun()

                elif _man_type == "force_camp":
                    _man_camp = st.text_input(
                        "Camp name / identifier:", placeholder="e.g. Dougie's camp",
                        key=_man_camp_key, label_visibility="collapsed"
                    )
                    if st.button("➕ Add Forced Camp", key=f"y9_ft_man_add_{_fi}_camp", use_container_width=True):
                        if _man_camp and _man_camp.strip():
                            _detail = f"{_fres['name']} → {_man_camp.strip()}"
                            _man_entries.append({"type": "force_camp", "detail": _detail})
                            st.session_state.y9_ft_manual_add[_man_email] = _man_entries
                            st.rerun()

                elif _man_type == "force_week":
                    _man_week = st.selectbox(
                        "Force to week:", [1, 2],
                        format_func=lambda w: f"Week {w}",
                        key=_man_week_key, label_visibility="collapsed"
                    )
                    if st.button("➕ Add Forced Week", key=f"y9_ft_man_add_{_fi}_week", use_container_width=True):
                        _wk = int(st.session_state.get(_man_week_key, 1))
                        st.session_state.y9_ft_new_force_week[_fres["name"]] = _wk
                        _detail = f"{_fres['name']} → Week {_wk}"
                        _man_entries.append({"type": "force_week", "detail": _detail})
                        st.session_state.y9_ft_manual_add[_man_email] = _man_entries
                        st.rerun()

                elif _man_type == "staff_note":
                    _man_note = st.text_area(
                        "Note text:", placeholder="Type the note to add to the staff export…",
                        key=_man_note_key, label_visibility="collapsed", height=80
                    )
                    if st.button("➕ Add Staff Note", key=f"y9_ft_man_add_{_fi}_note", use_container_width=True):
                        if _man_note and _man_note.strip():
                            _detail = _man_note.strip()
                            _man_entries.append({"type": "staff_note", "detail": _detail})
                            st.session_state.y9_ft_manual_add[_man_email] = _man_entries
                            st.rerun()

        # Error expander
        if _ft_errors:
            with st.expander(f"⚠️ {len(_ft_errors)} response(s) failed to parse — click to review"):
                for _ferr in _ft_errors:
                    st.warning(f"**{_ferr['name']}**: {_ferr.get('error_msg', 'Unknown error')}")
                    st.write(f"Original text: _{_ferr['text']}_")

        st.markdown("---")

        # ── Exports ───────────────────────────────────────────────────────────────────────
        st.subheader("📥 Exports")
        _exp_c1, _exp_c2 = st.columns(2)

        # ── Export 1: Combined Rules JSON ─────────────────────────────────────────────────
        with _exp_c1:
            st.markdown("**Export 1 — Combined Rules JSON**")
            st.caption(
                "Merges your accepted rules with any existing JSON. "
                "Load the result directly into the Y9 Journey Groups tool."
            )

            _ft_base = {
                "sep": [], "must": [], "force": {}, "force_week": {},
                "na": [], "na_details": [], "include_drafts": False, "weights": {}
            }
            if _ft_rules_upload is not None:
                try:
                    _ft_rules_upload.seek(0)
                    _loaded_rules = json.load(_ft_rules_upload)
                    for _jk in _ft_base:
                        if _jk in _loaded_rules:
                            _ft_base[_jk] = _loaded_rules[_jk]
                except Exception:
                    st.warning("⚠️ Could not parse the uploaded rules JSON — starting from empty.")

            # Append new must rules (deduplicated)
            _existing_must_set = {tuple(sorted(p)) for p in _ft_base["must"]}
            for _ma, _mb in st.session_state.y9_ft_new_must:
                _pk = tuple(sorted([_ma, _mb]))
                if _pk not in _existing_must_set:
                    _ft_base["must"].append([_ma, _mb])
                    _existing_must_set.add(_pk)

            # Append force-week overrides (deduplicated / override existing)
            for _fw_nm, _fw_wk in st.session_state.y9_ft_new_force_week.items():
                _ft_base["force_week"][_fw_nm] = _fw_wk

            st.download_button(
                "📥 Export Combined Rules JSON",
                data=json.dumps(_ft_base, indent=2).encode("utf-8"),
                file_name="y9_camp_rules_combined.json",
                mime="application/json",
                use_container_width=True,
                key="y9_ft_dl_json"
            )

            # Quick preview of new rules
            if st.session_state.y9_ft_new_must or st.session_state.y9_ft_new_force_week:
                with st.expander("Preview new rules being added"):
                    if st.session_state.y9_ft_new_must:
                        st.write("**New Must-Go pairs:**")
                        for _ma, _mb in st.session_state.y9_ft_new_must:
                            st.write(f"  • {_ma} + {_mb}")
                    if st.session_state.y9_ft_new_force_week:
                        st.write("**New Force-Week overrides:**")
                        for _fn, _fw in st.session_state.y9_ft_new_force_week.items():
                            st.write(f"  • {_fn} → Week {_fw}")

        # ── Export 2: Staff Notes Excel ───────────────────────────────────────────────────
        with _exp_c2:
            st.markdown("**Export 2 — Staff Notes Excel**")
            st.caption(
                "One row per student per category. "
                "Excludes pure friend requests and irrelevant responses."
            )

            _FT_NOTE_FIELDS = [
                ("medical_notes",    "Medical"),
                ("concern_notes",    "Concern"),
                ("permission_notes", "Permission"),
                ("tent_notes",       "Tent"),
                ("preference_notes", "Preference"),
                ("equipment_notes",  "Equipment"),
                ("logistics_notes",  "Logistics"),
                ("other_notes",      "Other"),
            ]

            _FT_CAT_STYLE = {
                "Medical":    ("A83232", "FFFFFF"),
                "Concern":    ("FF9933", "000000"),
                "Permission": ("FFCC00", "000000"),
                "Tent":       ("FF9900", "000000"),
                "Preference": ("C8F7C5", "000000"),
                "Equipment":  ("B5D4F4", "000000"),
                "Logistics":  ("CECBF6", "000000"),
                "Other":      ("D9D9D9", "000000"),
            }

            _ft_note_rows = []
            for _fres in _ft_res_all:
                if _fres["error"]: continue
                if _fres["llm"].get("is_irrelevant", False): continue
                _nm  = _fres["name"]
                _hse = _fres.get("house", "") or ""
                _wk  = _fres.get("week")
                _wk_str = f"Week {int(_wk)}" if (_wk and not pd.isna(_wk)) else "Unknown"

                for _fld, _cat in _FT_NOTE_FIELDS:
                    _raw_val = _fres["llm"].get(_fld)
                    if isinstance(_raw_val, list): _raw_val = " ".join(str(x) for x in _raw_val if x)
                    _val = (str(_raw_val) if _raw_val else "").strip()
                    if _val:
                        # Only include if the reviewer ticked it
                        _tick_key = (_fres["email"], _fld)
                        if st.session_state.y9_ft_staff_ticks.get(_tick_key, False):
                            _ft_note_rows.append({
                                "Student":           _nm,
                                "House":             _hse.title(),
                                "Week":              _wk_str,
                                "Category":          _cat,
                                "Summary":           _val,
                                "Original Response": _fres["text"],
                            })

                # Manual additions for this student
                # staff_note → Other; force_camp → Preference; pairing/force_week → Other/Logistics
                for _me in st.session_state.y9_ft_manual_add.get(_fres["email"], []):
                    if _me["type"] == "staff_note":
                        _ft_note_rows.append({
                            "Student":           _nm,
                            "House":             _hse.title(),
                            "Week":              _wk_str,
                            "Category":          "Other",
                            "Summary":           f"[Manual note] {_me['detail']}",
                            "Original Response": _fres["text"],
                        })
                    elif _me["type"] == "force_camp":
                        _ft_note_rows.append({
                            "Student":           _nm,
                            "House":             _hse.title(),
                            "Week":              _wk_str,
                            "Category":          "Preference",
                            "Summary":           f"[Forced camp] {_me['detail']}",
                            "Original Response": _fres["text"],
                        })
                    elif _me["type"] == "pairing":
                        _ft_note_rows.append({
                            "Student":           _nm,
                            "House":             _hse.title(),
                            "Week":              _wk_str,
                            "Category":          "Other",
                            "Summary":           f"[Must-Go pairing] {_me['detail']}",
                            "Original Response": _fres["text"],
                        })
                    elif _me["type"] == "force_week":
                        _ft_note_rows.append({
                            "Student":           _nm,
                            "House":             _hse.title(),
                            "Week":              _wk_str,
                            "Category":          "Logistics",
                            "Summary":           f"[Force-week override] {_me['detail']}",
                            "Original Response": _fres["text"],
                        })

            if _ft_note_rows:
                _ft_wb  = Workbook()
                _ft_ws  = _ft_wb.active
                _ft_ws.title = "Staff Notes"

                _ft_note_cols = ["Student", "House", "Week", "Category", "Summary", "Original Response"]

                for _ci, _col in enumerate(_ft_note_cols, 1):
                    _c = _ft_ws.cell(row=1, column=_ci, value=_col)
                    _c.fill = PatternFill("solid", fgColor="4472C4")
                    _c.font = Font(bold=True, color="FFFFFF")
                    _c.alignment = Alignment(horizontal="center", vertical="center")

                for _ri, _row in enumerate(_ft_note_rows, 2):
                    for _ci, _col in enumerate(_ft_note_cols, 1):
                        _c = _ft_ws.cell(row=_ri, column=_ci, value=_row[_col])
                        _c.alignment = Alignment(wrap_text=True, vertical="top")

                    _cat = _row["Category"]
                    if _cat in _FT_CAT_STYLE:
                        _fg, _fc = _FT_CAT_STYLE[_cat]
                        _cat_c = _ft_ws.cell(row=_ri, column=4)
                        _cat_c.fill = PatternFill("solid", fgColor=_fg)
                        _cat_c.font = Font(bold=True, color=_fc)
                        _cat_c.alignment = Alignment(horizontal="center", vertical="top")

                # Column widths
                _ft_col_widths = {
                    "Student": 22, "House": 12, "Week": 10,
                    "Category": 14, "Summary": 60, "Original Response": 60
                }
                for _ci, _col in enumerate(_ft_note_cols, 1):
                    _ft_ws.column_dimensions[get_column_letter(_ci)].width = _ft_col_widths.get(_col, 20)

                _ft_out = io.BytesIO()
                _ft_wb.save(_ft_out)
                _ft_out.seek(0)

                st.download_button(
                    "📥 Export Staff Notes (Excel)",
                    data=_ft_out,
                    file_name="Y9_Staff_Notes.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="y9_ft_dl_notes"
                )
                st.caption(f"{len(_ft_note_rows)} row(s) across {len(set(r['Student'] for r in _ft_note_rows))} student(s).")
            else:
                st.info("No staff notes to export — no relevant non-trivial content found yet.")

    elif not st.session_state.y9_ft_analysed:
        st.info(
            "Click **🔬 Analyse Free-Text Responses** above to run the AI over the survey. "
            "All processing happens on-device via Ollama — no data leaves your Mac."
        )